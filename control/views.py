from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Sum, Count, Case, When, Value, IntegerField
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
from django.db import IntegrityError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, timedelta, time
from decimal import Decimal

from control.utils import calcular_numerales_caja, get_inicio_ciclo
from .models import CierreTurno, CierreTurnoZona, CierreTurnoMovimiento, CierreTurnoPago, CierreTurnoDenominacion, CicloRecaudacion
from .forms import CierreTurnoForm, CierreTurnoZonaFormSet, CierreTurnoMovimientoFormSet, CierreTurnoPagoFormSet, CierreTurnoDenFormSet
from .models import CuadraturaDetalle
from django.views.generic.edit import UpdateView
from django.urls import reverse_lazy, reverse
from django.db import transaction
from django.db.models import Q
import calendar
from datetime import date
import json
import re
import base64
import io
from collections import Counter

import numpy as np
from PIL import Image, ImageOps
import pytesseract
from pytesseract import Output

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from .models import (
    Usuario, Sucursal, Zona, Maquina, Turno, LecturaMaquina,
    CuadraturaCajaDiaria, EncuadreCajaAdmin,
    AsignacionTurnoZona, AsignacionTurnoSlot,
    InformeRecaudacion, InformeRecaudacionLinea, InformeRecaudacionCaja,
    ControlLecturas, ControlLecturasLinea,
    ProgramacionRecaudacion,
    CuadraturaZona,
    RegistroSesion,
)

from .forms import (
    TurnoForm, LecturaMaquinaForm, SucursalForm, ZonaForm,
    MaquinaForm, UsuarioForm, UsuarioEditForm,
    CuadraturaCajaDiariaForm, EncuadreCajaAdminForm
)

from .decorators import (role_required, readonly_for,
    ROLES_DASHBOARD, ROLES_TURNO, ROLES_OPERACIONES, ROLES_REGISTRO,
    ROLES_RECAUDACION, ROLES_CONTROLES, ROLES_CONFIG_VER, ROLES_CONFIG_EDIT,
    ROLES_USUARIOS,ROLES_VER_USUARIOS ,ROLES_READONLY, ROLES_CUADRATURA, ROLES_CUADRATURA_DIARIA, ROLES_CUADRATURA_ZONA)
from django.shortcuts import render

def _estado_ord():
    """Anotación para ordenar máquinas: Mantenimiento → Retirada → Operativa."""
    return Case(
        When(estado="Mantenimiento", then=Value(0)),
        When(estado="Retirada",      then=Value(1)),
        default=Value(2),
        output_field=IntegerField(),
    )


def error_403(request, exception=None):
    return render(request, "errors/403.html", status=403)

#===========================
#Recaudación/Dia 0
#===========================
@login_required
@transaction.atomic
def iniciar_dia_0(request):
    if request.method != "POST":
        return redirect("control:recaudacion")

    sucursal_id = request.POST.get("sucursal_id")
    fecha_inicio_str = request.POST.get("fecha_inicio")

    if not sucursal_id:
        messages.error(request, "Debes seleccionar un local.")
        return redirect("control:recaudacion")

    if not fecha_inicio_str:
        messages.error(request, "Debes seleccionar una fecha de inicio del ciclo.")
        return redirect("control:recaudacion")

    try:
        from datetime import date
        fecha_inicio = date.fromisoformat(fecha_inicio_str)
    except ValueError:
        messages.error(request, "Fecha inválida.")
        return redirect("control:recaudacion")

    sucursal = get_object_or_404(Sucursal, id=sucursal_id)

    # 1) Actualizar contadores iniciales con la última lectura antes de la fecha de inicio
    maquinas = Maquina.objects.filter(sucursal=sucursal, is_active=True)
    actualizadas = 0
    sin_historial = 0

    for m in maquinas:
        ultimo = (LecturaMaquina.objects
                  .filter(maquina=m, fecha_trabajo__lte=fecha_inicio)
                  .order_by("-fecha_trabajo", "-id")
                  .first())

        if ultimo:
            m.contador_inicial_entrada = ultimo.entrada
            m.contador_inicial_salida = ultimo.salida
            m.save(update_fields=["contador_inicial_entrada", "contador_inicial_salida"])
            actualizadas += 1
        else:
            sin_historial += 1

    # 2) Guardar el ciclo con la fecha elegida y quién lo inició
    CicloRecaudacion.objects.update_or_create(
        sucursal=sucursal,
        defaults={
            "inicio_ciclo": fecha_inicio,
            "creado_por": request.user,
        },
    )

    messages.success(
        request,
        f"Día 0 iniciado para '{sucursal.nombre}' desde el {fecha_inicio.strftime('%d/%m/%Y')}. "
        f"Máquinas actualizadas: {actualizadas}. Sin historial: {sin_historial}."
    )
    return redirect("control:recaudacion")


@login_required

# ═══════════════════════════════════════════════════════════════
# INFORME DE RECAUDACIÓN
# ═══════════════════════════════════════════════════════════════

@login_required
def informe_recaudacion_list(request):
    """Lista de todos los informes de recaudación."""
    informes = (InformeRecaudacion.objects
                .select_related("sucursal", "creado_por")
                .all())
    sucursales = Sucursal.objects.filter(is_active=True).order_by("nombre")

    sucursal_id = request.GET.get("sucursal")
    if sucursal_id:
        informes = informes.filter(sucursal_id=sucursal_id)

    from django.core.paginator import Paginator
    paginator = Paginator(informes, 15)
    informes = paginator.get_page(request.GET.get("page", 1))
    return render(request, "recaudacion/informe_list.html", {
        "informes": informes,
        "page_obj": informes,
        "sucursales": sucursales,
        "sucursal_sel": sucursal_id,
    })


@login_required
def informe_recaudacion_detail(request, pk):
    """Detalle completo del informe — el "Excel" digital."""
    informe = get_object_or_404(
        InformeRecaudacion.objects.select_related("sucursal", "creado_por"),
        pk=pk
    )
    lineas = informe.lineas.select_related("zona").all()
    caja   = getattr(informe, "resumen_caja", None)

    # Agrupaciones para resúmenes
    from collections import defaultdict
    por_servidor = defaultdict(lambda: {"entrada": 0, "salida": 0, "total": 0, "cant": 0})
    por_juego    = defaultdict(lambda: {"entrada": 0, "salida": 0, "total": 0, "cant": 0})
    por_zona     = defaultdict(lambda: {"nombre": "", "entrada": 0, "salida": 0, "total": 0, "lineas": []})

    for l in lineas:
        s = l.servidor or "—"
        por_servidor[s]["entrada"] += l.parcial_entrada
        por_servidor[s]["salida"]  += l.parcial_salida
        por_servidor[s]["total"]   += l.total
        por_servidor[s]["cant"]    += 1

        j = l.juego or "—"
        por_juego[j]["entrada"] += l.parcial_entrada
        por_juego[j]["salida"]  += l.parcial_salida
        por_juego[j]["total"]   += l.total
        por_juego[j]["cant"]    += 1

        zn = l.zona_nombre or "Sin zona"
        por_zona[zn]["nombre"]  = zn
        por_zona[zn]["entrada"] += l.parcial_entrada
        por_zona[zn]["salida"]  += l.parcial_salida
        por_zona[zn]["total"]   += l.total
        por_zona[zn]["lineas"].append(l)

    # RTP para agrupaciones
    for d in por_servidor.values():
        d["rtp"] = round(d["salida"] / d["entrada"] * 100, 1) if d["entrada"] > 0 else 0
    for d in por_juego.values():
        d["rtp"] = round(d["salida"] / d["entrada"] * 100, 1) if d["entrada"] > 0 else 0
    for d in por_zona.values():
        d["rtp"] = round(d["salida"] / d["entrada"] * 100, 1) if d["entrada"] > 0 else 0

    return render(request, "recaudacion/informe_detail.html", {
        "informe": informe,
        "lineas": lineas,
        "caja": caja,
        "por_servidor": dict(sorted(por_servidor.items())),
        "por_juego":    dict(sorted(por_juego.items())),
        "por_zona":     dict(sorted(por_zona.items())),
    })


@login_required
def cerrar_ciclo_y_generar_informe(request):
    """
    Cierra el ciclo actual de una sucursal:
    1. Genera el InformeRecaudacion con snapshot de datos
    2. Actualiza contadores iniciales de las máquinas
    3. Actualiza CicloRecaudacion.inicio_ciclo a hoy
    """
    if request.method != "POST":
        return redirect("control:recaudacion")

    sucursal_id = request.POST.get("sucursal_id")
    sucursal    = get_object_or_404(Sucursal, pk=sucursal_id)

    ciclo = getattr(sucursal, "ciclo_recaudacion", None)
    if not ciclo:
        messages.error(request, "Esta sucursal no tiene un ciclo iniciado.")
        return redirect("control:recaudacion")

    fecha_inicio = ciclo.inicio_ciclo
    fecha_cierre = timezone.localdate()

    with transaction.atomic():
        # ── 1. Crear informe base ───────────────────────────────────────────
        informe = InformeRecaudacion.objects.create(
            sucursal=sucursal,
            fecha_inicio=fecha_inicio,
            fecha_cierre=fecha_cierre,
            creado_por=request.user,
        )

        # ── 2. Líneas por máquina ───────────────────────────────────────────
        maquinas = Maquina.objects.filter(sucursal=sucursal, is_active=True).select_related("zona").annotate(estado_ord=_estado_ord()).order_by("estado_ord", "zona__orden", "numero_maquina")
        bulk_lineas = []
        total_entrada = 0
        total_salida  = 0

        for maq in maquinas:
            # Último control del ciclo para esta máquina
            ultima_linea = (ControlLecturasLinea.objects
                .filter(
                    control__sucursal=sucursal,
                    control__fecha_trabajo__gte=fecha_inicio,
                    control__fecha_trabajo__lte=fecha_cierre,
                    maquina=maq,
                )
                .order_by("-control__fecha_trabajo", "-id")
                .first())

            if not ultima_linea:
                continue  # máquina sin lecturas en el ciclo

            hist_e_inicio = int(maq.contador_inicial_entrada or 0)
            hist_s_inicio = int(maq.contador_inicial_salida  or 0)
            hist_e_cierre = int(ultima_linea.entrada_historica or 0)
            hist_s_cierre = int(ultima_linea.salida_historica  or 0)

            parcial_e = hist_e_cierre - hist_e_inicio
            parcial_s = hist_s_cierre - hist_s_inicio
            total_maq = parcial_e - parcial_s

            total_entrada += parcial_e
            total_salida  += parcial_s

            bulk_lineas.append(InformeRecaudacionLinea(
                informe=informe,
                zona=maq.zona,
                zona_nombre=maq.zona.nombre if maq.zona else "",
                numero_maquina=maq.numero_maquina,
                codigo_interno=maq.codigo_interno or "",
                servidor=maq.servidor or "",
                juego=maq.nombre_juego or "",
                hist_entrada_inicio=hist_e_inicio,
                hist_salida_inicio=hist_s_inicio,
                hist_entrada_cierre=hist_e_cierre,
                hist_salida_cierre=hist_s_cierre,
                parcial_entrada=parcial_e,
                parcial_salida=parcial_s,
                total=total_maq,
            ))

        InformeRecaudacionLinea.objects.bulk_create(bulk_lineas)

        # ── 3. Resumen caja (acumulados del ciclo) ─────────────────────────
        # Tomamos los campos _acum del ÚLTIMO día del ciclo (ya son acumulativos)
        ultima_cuad = (CuadraturaCajaDiaria.objects
            .filter(sucursal=sucursal, fecha__gte=fecha_inicio, fecha__lte=fecha_cierre)
            .order_by("-fecha", "-creado_el")
            .first())

        # Sumar retiro_diario de todos los días del ciclo
        retiros_total = (CuadraturaCajaDiaria.objects
            .filter(sucursal=sucursal, fecha__gte=fecha_inicio, fecha__lte=fecha_cierre)
            .aggregate(s=Sum("retiro_diario"))["s"] or 0)

        if ultima_cuad:
            InformeRecaudacionCaja.objects.create(
                informe=informe,
                numeral=int(ultima_cuad.numeral_acumulado or 0),
                prestamos=int(ultima_cuad.prestamos_acum or 0),
                retiros=int(retiros_total),
                redbank=int(ultima_cuad.redbank_acum or 0),
                transferencias=int(ultima_cuad.transfer_acum or 0),
                sueldo_extra=int(ultima_cuad.sueldo_b_acum or 0),
                sorteos=int(ultima_cuad.sorteos_acum or 0),
                gastos=int(ultima_cuad.gastos_acum or 0),
                regalos=int(ultima_cuad.regalos_acum or 0),
                jugados=int(ultima_cuad.jugados_acum or 0),
                otros=int((ultima_cuad.otros_1_acum or 0) + (ultima_cuad.otros_2_acum or 0)),
                billetes_malos=int(ultima_cuad.ef_billetes_malos or 0),
                descuadre=int(ultima_cuad.descuadre_acum or 0),
                total_caja=int(ultima_cuad.total_efectivo or 0),
            )
        else:
            InformeRecaudacionCaja.objects.create(informe=informe)

        # ── 4. Actualizar totales en el informe ────────────────────────────
        informe.total_numeral = int(ultima_cuad.numeral_acumulado or 0) if ultima_cuad else 0
        informe.total_entrada = total_entrada
        informe.total_salida  = total_salida
        informe.save(update_fields=["total_numeral", "total_entrada", "total_salida"])

        # ── 5. Actualizar contadores iniciales de máquinas ─────────────────
        for linea in bulk_lineas:
            Maquina.objects.filter(
                sucursal=sucursal,
                numero_maquina=linea.numero_maquina,
                zona=linea.zona,
            ).update(
                contador_inicial_entrada=linea.hist_entrada_cierre,
                contador_inicial_salida=linea.hist_salida_cierre,
            )

        # ── 6. Actualizar inicio_ciclo a hoy ───────────────────────────────
        CicloRecaudacion.objects.filter(sucursal=sucursal).update(
            inicio_ciclo=fecha_cierre,
            creado_por=request.user,
        )

    messages.success(request, f"Ciclo cerrado. Informe de recaudación generado para {sucursal.nombre}.")
    return redirect("control:informe_recaudacion_detail", pk=informe.pk)


@login_required
@role_required(*ROLES_RECAUDACION)
def recaudacion_view(request):
    sucursales = Sucursal.objects.filter(is_active=True).order_by("nombre")
    hoy = timezone.localdate()
    ciclos = (CicloRecaudacion.objects
              .select_related("sucursal", "creado_por")
              .order_by("sucursal__nombre"))
    return render(request, "recaudacion/recaudacion.html", {
        "sucursales": sucursales,
        "hoy": hoy,
        "ciclos": ciclos,
        "tiene_programaciones": ProgramacionRecaudacion.objects.filter(activa=True).exists(),
    })





# ==========================
# ROLES
# ==========================

def is_admin(user):
    return user.is_authenticated and getattr(user, "role", "") == "admin"

def is_usuario(user):
    return user.is_authenticated and getattr(user, "role", "") == "usuario"


# ==========================
# AUTH
# ==========================

# ═══════════════════════════════════════════════════════════════
# PROGRAMACIÓN AUTOMÁTICA DE RECAUDACIÓN
# ═══════════════════════════════════════════════════════════════

def _ejecutar_recaudacion_programada(sucursal, user):
    """
    Genera el informe de recaudación automáticamente si corresponde.
    Igual que cerrar_ciclo_y_generar_informe pero sin request.
    """
    ciclo = getattr(sucursal, "ciclo_recaudacion", None)
    if not ciclo:
        return None

    fecha_inicio = ciclo.inicio_ciclo
    fecha_cierre = timezone.localdate()

    # No cerrar si ya existe un informe para este cierre
    if InformeRecaudacion.objects.filter(sucursal=sucursal, fecha_cierre=fecha_cierre).exists():
        return None

    with transaction.atomic():
        informe = InformeRecaudacion.objects.create(
            sucursal=sucursal,
            fecha_inicio=fecha_inicio,
            fecha_cierre=fecha_cierre,
            creado_por=user,
        )

        maquinas = Maquina.objects.filter(sucursal=sucursal, is_active=True).select_related("zona").annotate(estado_ord=_estado_ord()).order_by("estado_ord", "zona__orden", "numero_maquina")
        bulk_lineas = []
        total_entrada = total_salida = 0

        for maq in maquinas:
            ultima_linea = (ControlLecturasLinea.objects
                .filter(
                    control__sucursal=sucursal,
                    control__fecha_trabajo__gte=fecha_inicio,
                    control__fecha_trabajo__lte=fecha_cierre,
                    maquina=maq,
                )
                .order_by("-control__fecha_trabajo", "-id")
                .first())
            if not ultima_linea:
                continue

            hist_e_inicio = int(maq.contador_inicial_entrada or 0)
            hist_s_inicio = int(maq.contador_inicial_salida  or 0)
            hist_e_cierre = int(ultima_linea.entrada_historica or 0)
            hist_s_cierre = int(ultima_linea.salida_historica  or 0)
            parcial_e = hist_e_cierre - hist_e_inicio
            parcial_s = hist_s_cierre - hist_s_inicio
            total_maq = parcial_e - parcial_s
            total_entrada += parcial_e
            total_salida  += parcial_s

            bulk_lineas.append(InformeRecaudacionLinea(
                informe=informe,
                zona=maq.zona,
                zona_nombre=maq.zona.nombre if maq.zona else "",
                numero_maquina=maq.numero_maquina,
                codigo_interno=maq.codigo_interno or "",
                servidor=maq.servidor or "",
                juego=maq.nombre_juego or "",
                hist_entrada_inicio=hist_e_inicio,
                hist_salida_inicio=hist_s_inicio,
                hist_entrada_cierre=hist_e_cierre,
                hist_salida_cierre=hist_s_cierre,
                parcial_entrada=parcial_e,
                parcial_salida=parcial_s,
                total=total_maq,
            ))

        InformeRecaudacionLinea.objects.bulk_create(bulk_lineas)

        ultima_cuad = (CuadraturaCajaDiaria.objects
            .filter(sucursal=sucursal, fecha__gte=fecha_inicio, fecha__lte=fecha_cierre)
            .order_by("-fecha", "-creado_el").first())
        retiros_total = (CuadraturaCajaDiaria.objects
            .filter(sucursal=sucursal, fecha__gte=fecha_inicio, fecha__lte=fecha_cierre)
            .aggregate(s=Sum("retiro_diario"))["s"] or 0)

        if ultima_cuad:
            InformeRecaudacionCaja.objects.create(
                informe=informe,
                numeral=int(ultima_cuad.numeral_acumulado or 0),
                prestamos=int(ultima_cuad.prestamos_acum or 0),
                retiros=int(retiros_total),
                redbank=int(ultima_cuad.redbank_acum or 0),
                transferencias=int(ultima_cuad.transfer_acum or 0),
                sueldo_extra=int(ultima_cuad.sueldo_b_acum or 0),
                sorteos=int(ultima_cuad.sorteos_acum or 0),
                gastos=int(ultima_cuad.gastos_acum or 0),
                regalos=int(ultima_cuad.regalos_acum or 0),
                jugados=int(ultima_cuad.jugados_acum or 0),
                otros=int((ultima_cuad.otros_1_acum or 0) + (ultima_cuad.otros_2_acum or 0)),
                billetes_malos=int(ultima_cuad.ef_billetes_malos or 0),
                descuadre=int(ultima_cuad.descuadre_acum or 0),
                total_caja=int(ultima_cuad.total_efectivo or 0),
            )
        else:
            InformeRecaudacionCaja.objects.create(informe=informe)

        informe.total_numeral = int(ultima_cuad.numeral_acumulado or 0) if ultima_cuad else 0
        informe.total_entrada = total_entrada
        informe.total_salida  = total_salida
        informe.save(update_fields=["total_numeral", "total_entrada", "total_salida"])

        for linea in bulk_lineas:
            Maquina.objects.filter(
                sucursal=sucursal,
                numero_maquina=linea.numero_maquina,
                zona=linea.zona,
            ).update(
                contador_inicial_entrada=linea.hist_entrada_cierre,
                contador_inicial_salida=linea.hist_salida_cierre,
            )

        CicloRecaudacion.objects.filter(sucursal=sucursal).update(
            inicio_ciclo=fecha_cierre,
            creado_por=user,
        )

    return informe


def _chequear_programaciones(user):
    """
    Llamado en el login. Para cada sucursal con programación activa,
    verifica si hoy es el día del mes configurado y ya pasó la hora → ejecuta cierre.
    Devuelve lista de informes recién generados (para mostrar notificación).
    """
    ahora      = timezone.localtime(timezone.now())
    hoy        = ahora.date()
    hora_ahora = ahora.time()
    dia_hoy    = hoy.day  # 1-31

    nuevos = []
    for prog in ProgramacionRecaudacion.objects.filter(activa=True).select_related("sucursal"):
        # Ajuste para meses cortos: si dia_del_mes=31 y el mes tiene 28, ejecutar el último día
        import calendar
        ultimo_dia = calendar.monthrange(hoy.year, hoy.month)[1]
        dia_efectivo = min(prog.dia_del_mes, ultimo_dia)

        if dia_efectivo != dia_hoy:
            continue
        if hora_ahora < prog.hora:
            continue
        informe = _ejecutar_recaudacion_programada(prog.sucursal, user)
        if informe:
            nuevos.append(informe)

    return nuevos


@login_required
def dismiss_notificacion_recaudacion(request, pk):
    """El primer usuario que la ve la marca como consumida."""
    if request.method == "POST":
        InformeRecaudacion.objects.filter(pk=pk, notificacion_consumida=False).update(
            notificacion_consumida=True
        )
        # Limpiar sesión
        ids = request.session.get("notif_recaudacion_ids", [])
        ids = [i for i in ids if i != pk]
        if ids:
            request.session["notif_recaudacion_ids"] = ids
        else:
            request.session.pop("notif_recaudacion_ids", None)
        request.session.modified = True
    next_url = request.POST.get("next", "/")
    return redirect(next_url)


@login_required
def programacion_recaudacion_view(request):
    """Configurar programación automática por sucursal."""
    sucursales = Sucursal.objects.filter(is_active=True).order_by("nombre")

    if request.method == "POST":
        sucursal_id = request.POST.get("sucursal_id")
        sucursal    = get_object_or_404(Sucursal, pk=sucursal_id)
        activa      = request.POST.get("activa") == "1"
        dia         = int(request.POST.get("dia_del_mes", 1))
        dia         = max(1, min(31, dia))  # clamp 1-31
        hora_str    = request.POST.get("hora", "23:00")

        from datetime import time as dtime
        h, mi = map(int, hora_str.split(":"))

        ProgramacionRecaudacion.objects.update_or_create(
            sucursal=sucursal,
            defaults={
                "activa": activa,
                "dia_del_mes": dia,
                "hora": dtime(h, mi),
                "creado_por": request.user,
            }
        )
        messages.success(request, f"Programación guardada para {sucursal.nombre}.")
        return redirect("control:programacion_recaudacion")

    q = request.GET.get("q", "").strip()
    if q:
        sucursales = sucursales.filter(nombre__icontains=q)

    programaciones = {p.sucursal_id: p for p in ProgramacionRecaudacion.objects.all()}

    sucursales_config = [
        {"sucursal": s, "prog": programaciones.get(s.id)}
        for s in sucursales
    ]

    return render(request, "recaudacion/programacion.html", {
        "sucursales_config": sucursales_config,
        "q": q,
    })


def _get_tipo_turno_por_hora():
    """
    Devuelve el tipo de turno según la hora actual:
      Mañana: 08:00 - 11:59
      Tarde:  12:00 - 19:59
      Noche:  20:00 - 07:59 (cruza medianoche)
    """
    from datetime import datetime
    hora = datetime.now().hour
    if 8 <= hora <= 11:
        return "Mañana"
    elif 12 <= hora <= 19:
        return "Tarde"
    else:
        return "Noche"


def _redirect_por_rol(role, user):
    """Retorna la URL de redirección según el rol del usuario."""
    if user.role in ('admin',):
        return redirect("control:dashboard")
    if user.role == 'gerente' or role == 'supervisor':
        return redirect("control:maquinas_list")
    if user.role == 'tecnico':
        return redirect("control:maquinas_list")
    if user.role == 'asistente':
        turno = Turno.objects.filter(usuario=user, estado="Abierto").first()
        if turno:
            return redirect("control:cuadratura_zona", turno_id=turno.id)
        return redirect("control:turno")
    # encargado, asistente
    return redirect("control:turno")


def login_view(request):
    if request.user.is_authenticated:
        return _redirect_por_rol(request.user.role, request.user)

    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")

        # ── Rate limiting: máx 10 intentos fallidos por IP en 10 minutos ────
        from django.core.cache import cache as _login_cache
        ip_key = "login_fail_" + (request.META.get("HTTP_X_FORWARDED_FOR", request.META.get("REMOTE_ADDR", "unknown")) or "unknown").split(",")[0].strip()
        intentos = _login_cache.get(ip_key, 0)
        if intentos >= 10:
            messages.error(request, "Demasiados intentos fallidos. Espera 10 minutos.")
            return render(request, "login.html", {"alerta_sesion": ""})

        user = authenticate(request, username=username, password=password)
        if user is not None:
            _login_cache.delete(ip_key)  # reset counter on success
            login(request, user)
            # Chequear programaciones automáticas
            nuevos_informes = _chequear_programaciones(user)
            # Guardar en sesión los IDs de informes no consumidos (solo el primero los ve)
            pendientes = list(
                InformeRecaudacion.objects
                .filter(notificacion_consumida=False)
                .values_list("id", flat=True)
                .order_by("-creado_en")[:10]
            )
            if pendientes:
                request.session["notif_recaudacion_ids"] = pendientes
            # Crear registro de sesión (sucursal/turno se actualizan después)
            ip = request.META.get('HTTP_X_FORWARDED_FOR', request.META.get('REMOTE_ADDR', ''))
            if ip and ',' in ip:
                ip = ip.split(',')[0].strip()
            # Cerrar sesiones huérfanas del mismo usuario (browser cerrado sin logout)
            RegistroSesion.objects.filter(
                usuario=user,
                hora_cierre__isnull=True
            ).update(
                hora_cierre=timezone.now(),
                motivo_cierre="nueva_sesion",
            )

            registro = RegistroSesion.objects.create(
                usuario=user,
                tipo_usuario=user.role,
            )
            request.session['registro_sesion_id'] = registro.pk

            redir = _manejar_sucursal_post_login(request, user)

            # Si no hubo redirección, actualizar sucursal y turno ahora
            if not redir:
                sucursal_id = request.session.get('sucursal_activa_id')
                turno_obj = Turno.objects.filter(usuario=user, estado="Abierto").first()
                RegistroSesion.objects.filter(pk=registro.pk).update(
                    sucursal_id=sucursal_id or None,
                    turno=turno_obj,
                )
                return _redirect_por_rol(user.role, user)
            return redir
        else:
            # Incrementar contador de intentos fallidos
            from django.core.cache import cache as _login_cache
            ip_key2 = "login_fail_" + (request.META.get("HTTP_X_FORWARDED_FOR", request.META.get("REMOTE_ADDR", "unknown")) or "unknown").split(",")[0].strip()
            _login_cache.set(ip_key2, _login_cache.get(ip_key2, 0) + 1, timeout=600)
            messages.error(request, "Usuario o contraseña incorrectos.")

    # Detectar si el usuario fue desconectado forzosamente (sin_asignacion)
    alerta_sesion = ""
    # Detectar cierre forzado: cuando llega al login con ?next= (redirigido por Django)
    # y hay un flag en caché con su uid (seteado al eliminar la sesión)
    from django.core.cache import cache as _cache
    uid_alerta = request.GET.get("uid", "")
    if uid_alerta and _cache.get(f"forzado_sin_asig_{uid_alerta}"):
        alerta_sesion = "Tu sesión fue cerrada porque ya no tienes ninguna asignación activa en el turno."
        _cache.delete(f"forzado_sin_asig_{uid_alerta}")
        try:
            RegistroSesion.objects.filter(
                usuario_id=int(uid_alerta), motivo_cierre="sin_asignacion"
            ).update(motivo_cierre=None)
        except Exception:
            pass
    return render(request, "login.html", {"alerta_sesion": alerta_sesion})


def _verificar_asignacion_asistente(user, sucursal):
    """
    Verifica si el asistente tiene al menos una asignación activa
    (zona, redbank o servicio) en el turno abierto del encargado de la sucursal.
    """
    turno = Turno.objects.filter(sucursal=sucursal, estado="Abierto", usuario__role="encargado").first()
    if not turno:
        return False
    tiene_zona = AsignacionTurnoZona.objects.filter(turno=turno, usuario=user).exists()
    tiene_slot = AsignacionTurnoSlot.objects.filter(turno=turno, usuario=user).exists()
    return tiene_zona or tiene_slot


def _manejar_sucursal_post_login(request, user):
    """
    Para encargado/asistente:
    - Encargado: 1 suc → entra directo | 2+ → pantalla selección
    - Asistente: verifica asignación activa en turno del encargado de cada sucursal
        · ninguna con asignación → error + logout
        · solo una con asignación → entra directo
        · varias con asignación  → pantalla selección solo con esas
    Otros roles → retorna None (sin efecto).
    """
    from django.contrib.auth import logout as auth_logout

    ROLES_CON_SELECCION = ('encargado', 'asistente')
    if user.role not in ROLES_CON_SELECCION:
        return None

    sucursales = list(user.sucursales.filter(is_active=True))

    if not sucursales:
        return None

    if user.role == 'asistente':
        sucursales_con_asignacion = [s for s in sucursales if _verificar_asignacion_asistente(user, s)]

        if not sucursales_con_asignacion:
            auth_logout(request)
            messages.error(
                request,
                "No puedes iniciar sesión porque no tienes ninguna asignación "
                "(zona, redbank o servicio) en ningún turno activo."
            )
            return redirect("control:login")

        if len(sucursales_con_asignacion) == 1:
            sucursal = sucursales_con_asignacion[0]
            request.session['sucursal_activa_id'] = sucursal.pk
            _auto_iniciar_turno(request, user, sucursal)
            return None

        request.session['sucursales_asistente_ids'] = [s.pk for s in sucursales_con_asignacion]
        return redirect("control:seleccionar_sucursal")

    # ── Encargado ──
    if len(sucursales) == 1:
        request.session['sucursal_activa_id'] = sucursales[0].pk
        _auto_iniciar_turno(request, user, sucursales[0])
        return None
    else:
        return redirect("control:seleccionar_sucursal")


@login_required
def seleccionar_sucursal_view(request):
    """
    Pantalla de selección de sucursal.
    - Encargado: ve todas sus sucursales
    - Asistente: solo llega aquí si tiene asignación en 2+ sucursales
    """
    ROLES_CON_SELECCION = ('encargado', 'asistente')
    user = request.user

    if user.role not in ROLES_CON_SELECCION:
        return _redirect_por_rol(user.role, user)

    if user.role == 'asistente':
        from django.contrib.auth import logout as auth_logout
        ids = request.session.get('sucursales_asistente_ids', [])
        if not ids:
            auth_logout(request)
            messages.error(request, "No tienes ninguna asignación activa.")
            return redirect("control:login")
        sucursales = user.sucursales.filter(pk__in=ids, is_active=True).order_by('nombre')
    else:
        sucursales = user.sucursales.filter(is_active=True).order_by('nombre')

    if sucursales.count() == 1:
        sucursal = sucursales.first()
        request.session['sucursal_activa_id'] = sucursal.pk
        request.session.pop('sucursales_asistente_ids', None)
        if user.role in ('encargado', 'asistente'):
            _auto_iniciar_turno(request, user, sucursal)
        return _redirect_por_rol(user.role, user)

    if request.method == 'POST':
        sucursal_id = request.POST.get('sucursal_id')
        if sucursal_id and sucursales.filter(pk=sucursal_id).exists():
            sucursal = sucursales.get(pk=sucursal_id)
            request.session['sucursal_activa_id'] = int(sucursal_id)
            request.session.pop('sucursales_asistente_ids', None)
            if user.role in ('encargado', 'asistente'):
                _auto_iniciar_turno(request, user, sucursal)
            return _redirect_por_rol(user.role, user)
        else:
            messages.error(request, 'Debes seleccionar una sucursal válida.')

    return render(request, 'seleccionar_sucursal.html', {'sucursales': sucursales})


def _auto_iniciar_turno(request, user, sucursal):
    """
    Crea el turno automáticamente para el encargado según la hora actual.
    Si ya tiene un turno abierto, no crea uno nuevo.
    """
    from datetime import date as date_today, datetime, timedelta

    if Turno.objects.filter(usuario=user, estado="Abierto").first():
        return  # Ya tiene turno abierto, no crear otro

    tipo_turno = _get_tipo_turno_por_hora()
    hoy = date_today.today()

    # Turno noche que cruza medianoche (00:00-07:59) → fecha del día anterior
    hora = datetime.now().hour
    if tipo_turno == "Noche" and 0 <= hora <= 7:
        hoy = hoy - timedelta(days=1)

    turno = Turno(
        sucursal=sucursal,
        usuario=user,
        fecha=hoy,
        tipo_turno=tipo_turno,
        estado="Abierto",
    )
    try:
        turno.full_clean()
        turno.save()
        messages.success(
            request,
            f"Turno {tipo_turno} iniciado automáticamente en {sucursal.nombre}."
        )
        # Actualizar registro de sesión con sucursal y turno
        reg_id = request.session.get('registro_sesion_id')
        if reg_id:
            RegistroSesion.objects.filter(pk=reg_id).update(
                sucursal=sucursal,
                turno=turno,
            )
    except Exception as e:
        messages.warning(request, f"No se pudo iniciar el turno: {e}")


@login_required
def logout_view(request):
    user = request.user

    # Solo el encargado pasa por la pantalla de advertencia/confirmación
    if user.role == 'encargado':
        turno_abierto = Turno.objects.filter(usuario=user, estado="Abierto").first()

        if not request.GET.get("confirmar"):
            # Siempre mostrar pantalla de confirmación al encargado
            maquinas_total = 0
            lecturas_registradas = 0
            maquinas_sin_lectura = 0

            if turno_abierto:
                maquinas_total = Maquina.objects.filter(
                    sucursal=turno_abierto.sucursal, is_active=True, zona__is_active=True
                ).count()
                lecturas_registradas = LecturaMaquina.objects.filter(turno=turno_abierto).count()
                maquinas_sin_lectura = maquinas_total - lecturas_registradas

            return render(request, "logout_advertencia.html", {
                "turno": turno_abierto,
                "maquinas_sin_lectura": maquinas_sin_lectura,
                "maquinas_total": maquinas_total,
                "lecturas_registradas": lecturas_registradas,
                "completo": maquinas_sin_lectura == 0,
            })

        # Confirmó → cerrar turno si está abierto
        if turno_abierto:
            total = LecturaMaquina.objects.filter(turno=turno_abierto).aggregate(
                total_sum=Sum("total")
            )["total_sum"] or 0
            turno_abierto.estado = "Cerrado"
            turno_abierto.total_cierre = Decimal(str(total))
            turno_abierto.save()

    # Cerrar registro de sesión
    from django.utils import timezone as tz
    reg_id = request.session.get('registro_sesion_id')
    if reg_id:
        RegistroSesion.objects.filter(pk=reg_id, hora_cierre__isnull=True).update(
            hora_cierre=tz.now()
        )

    logout(request)
    return redirect("control:login")


# Ruta a Tesseract (Windows)
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

@login_required
def ajax_cuadratura_mensual_data(request):
    sucursal_id = request.GET.get("sucursal_id")
    mes = request.GET.get("mes")  # "2026-01"

    if not sucursal_id or not mes:
        return JsonResponse({"ok": False, "error": "Faltan parámetros"}, status=400)

    y, m = map(int, mes.split("-"))
    last_day = calendar.monthrange(y, m)[1]
    fecha_desde = date(y, m, 1)
    fecha_hasta = date(y, m, last_day)

    qs = (
        LecturaMaquina.objects
        .filter(
            zona__sucursal_id=sucursal_id,
            fecha_registro__range=(fecha_desde, fecha_hasta),  
        )
        .values("zona_id", "zona__nombre")
        .annotate(
            entrada=Sum("entrada"),
            salida=Sum("salida"),
            total=Sum("total"),
        )
        .order_by("zona__nombre")
    )

    data = []
    for row in qs:
        data.append({
            "zona_id": row["zona_id"],
            "zona_nombre": row["zona__nombre"],
            "entrada": row["entrada"] or 0,
            "salida": row["salida"] or 0,
            "total": row["total"] or 0,
        })

    return JsonResponse({"ok": True, "zonas": data})
# ============================================
# CUADRATURA CAJA DIARIA (ATENDEDORAS)
# ============================================
@login_required
def cuadratura_create(request, turno_id):
    turno = get_object_or_404(Turno, id=turno_id)

    # Aquí renderizas el TEMPLATE DE "columna roja"
    # OJO: debe existir templates/cuadratura/create_turno.html  (o create.html si así lo llamas)
    return render(request, "cuadratura/create_turno.html", {
        "turno": turno,
        "sucursal": turno.sucursal,
        # "zonas": ...
        # "lecturas": ...
    })

@login_required
def cuadratura_list(request):
    return render(request, "cuadratura/list.html")

@login_required
def cuadratura_detail(request, pk):
    return render(request, "cuadratura/detail.html", {"pk": pk})

def _parse_detalles_from_post(post):
    """
    Lee inputs hidden con formato:
      detalles[0][tipo], detalles[0][nombre], detalles[0][monto]
    Retorna lista: [{"tipo":..., "nombre":..., "monto": int}, ...]
    """
    detalles = []
    idx = 0
    while True:
        k_tipo = f"detalles[{idx}][tipo]"
        k_nom = f"detalles[{idx}][nombre]"
        k_mon = f"detalles[{idx}][monto]"

        if k_tipo not in post and k_nom not in post and k_mon not in post:
            break

        tipo = (post.get(k_tipo) or "").strip().upper()
        nombre = (post.get(k_nom) or "").strip()
        monto_raw = (post.get(k_mon) or "").strip()

        # limpia puntos y basura
        monto_raw = monto_raw.replace(".", "")
        monto_raw = "".join(ch for ch in monto_raw if ch.isdigit() or ch == "-" )

        try:
            monto = int(monto_raw) if monto_raw not in ("", "-") else 0
        except ValueError:
            monto = 0

        # evita filas vacías
        if tipo and (nombre or monto != 0):
            detalles.append({"tipo": tipo, "nombre": nombre, "monto": monto})

        idx += 1

    return detalles


def _sum_tipo(cuadratura, tipo: str) -> int:
    return (
        CuadraturaDetalle.objects
        .filter(cuadratura=cuadratura, tipo=tipo)
        .aggregate(s=Sum("monto"))["s"]
        or 0
    )


def _upsert_cuadratura_diaria(*, sucursal, fecha, defaults):
    """
    Detecta si ya existe una cuadratura para (sucursal, fecha) y la actualiza.
    Si no existe, la crea.
    """
    obj = CuadraturaCajaDiaria.objects.filter(sucursal=sucursal, fecha=fecha).order_by("-creado_el").first()
    if obj:
        for k, v in defaults.items():
            setattr(obj, k, v)
        return obj, False
    return CuadraturaCajaDiaria(sucursal=sucursal, fecha=fecha, **defaults), True


def _recalcular_totales(cuadratura):
    """
    Implementa la lógica del bloc de notas.
    - Día 1 del ciclo: base anterior = caja_inicial del local
    - Día siguiente: base anterior = desglose_efectivo_total del día anterior (última cuadratura dentro del ciclo)
    - base_total = base_anterior + numeral_dia + prestamos
    - restas principales: sorteos, gastos, sueldos, regalos, jugados
    - restas adicionales: redbank, transfer
    - total_calculado = base_total - (restas)
    - descuadre = desglose_total - total_calculado
    - ganancia: (definimos UNA regla, ver abajo)
    - total_efectivo = desglose_total - retiro_diario
    """

    # 1) Numerales (backend es la verdad)
    numeral_dia, numeral_acum = calcular_numerales_caja(cuadratura.sucursal, cuadratura.fecha)
    cuadratura.numeral_dia = numeral_dia
    cuadratura.numeral_acumulado = numeral_acum

    # 2) Caja anterior dentro del ciclo (si es primer día del ciclo → caja_inicial)
    base_anterior, _ = _caja_anterior_en_ciclo(cuadratura.sucursal, cuadratura.fecha)

    # (si quieres guardar caja anterior en el modelo, aquí NO tienes campo.
    #  Solo lo usamos para cálculo)

    # 3) Base total
    prestamos = (cuadratura.prestamos or 0)  # puede ser negativo/positivo
    base_total = (base_anterior or 0) + (cuadratura.numeral_dia or 0) + prestamos

    # 4) Restas
    restas_principales = (
        (cuadratura.sorteos_dia or 0)
        + (cuadratura.gastos_dia or 0)
        + (cuadratura.sueldo_b_dia or 0)
        + (cuadratura.regalos_dia or 0)
        + (cuadratura.jugados_dia or 0)
    )

    restas_adicionales = (cuadratura.redbank_dia or 0) + (cuadratura.transfer_dia or 0)

    total_calculado = (base_total or 0) - (restas_principales or 0) - (restas_adicionales or 0)

    # 5) Desglose real contado
    # desglose_efectivo_total es un @property, se accede directamente
    desglose_total = cuadratura.desglose_efectivo_total or 0

    # 6) Descuadre (lo que conté vs lo que debería dar)
    cuadratura.descuadre_dia = desglose_total - (total_calculado or 0)

    # 7) Prestamos acumulados = prestamos de dias anteriores + dia actual
    prestamos_dia = int(cuadratura.prestamos or 0)
    _, prestamos_acum_ant = _caja_anterior_en_ciclo(cuadratura.sucursal, cuadratura.fecha)
    prestamos_acum = prestamos_acum_ant + prestamos_dia
    cuadratura.prestamos_acum = prestamos_acum

    # 8) Ganancia = desglose efectivo - caja inicial - prestamos acumulados
    caja_inicial_sucursal = int(cuadratura.sucursal.caja_inicial or 0)
    cuadratura.ganancia = desglose_total - caja_inicial_sucursal - prestamos_acum

    # 8) Total efectivo (contado menos retiro)
    cuadratura.total_efectivo = desglose_total - (cuadratura.retiro_diario or 0)

    return cuadratura

@login_required
@role_required(*ROLES_CUADRATURA_DIARIA)
def cuadratura_diaria_create(request):
    if request.method == "POST":
        form = CuadraturaCajaDiariaForm(request.POST)

        if form.is_valid():
            sucursal = form.cleaned_data.get("sucursal")
            fecha = form.cleaned_data["fecha"]

            if not sucursal:
                messages.error(request, "Debes seleccionar un Local (Sucursal).")
                return render(request, "cuadratura_diaria/create.html", {
                    "form": form,
                    "sucursales": Sucursal.objects.filter(is_active=True).order_by("nombre"),
                    "mes_default": timezone.localdate().strftime("%Y-%m"),
                })

            with transaction.atomic():
                cuadratura = form.save(commit=False)

                cuadratura.usuario = request.user
                cuadratura.sucursal = sucursal

                # Asignar campos ef_* manualmente
                for field in ['ef_20000', 'ef_10000', 'ef_5000', 'ef_2000', 'ef_1000', 'ef_monedas', 'ef_billetes_malos']:
                    setattr(cuadratura, field, form.cleaned_data.get(field, 0))

                # desglose_efectivo_total es un @property, se calcula automáticamente
                # total_efectivo = lo que queda en caja después del retiro diario
                cuadratura.total_efectivo = cuadratura.desglose_efectivo_total - (cuadratura.retiro_diario or 0)

                # timestamps (por tu tema MySQL)
                if not cuadratura.creado_el:
                    cuadratura.creado_el = timezone.now()
                cuadratura.actualizado_el = timezone.now()

                cuadratura.save()

                # Detalles
                detalles_list = _parse_detalles_from_post(request.POST)
                CuadraturaDetalle.objects.filter(cuadratura=cuadratura).delete()

                for d in detalles_list:
                    CuadraturaDetalle.objects.create(
                        cuadratura=cuadratura,
                        tipo=d["tipo"],
                        nombre=d["nombre"],
                        monto=d["monto"],
                        detalle="",
                    )

                # Re-sumar desde BD
                cuadratura.gastos_dia   = _sum_tipo(cuadratura, "GASTOS")
                cuadratura.sueldo_b_dia = _sum_tipo(cuadratura, "SUELDOS")
                cuadratura.regalos_dia  = _sum_tipo(cuadratura, "REGALOS")
                cuadratura.taxi_dia     = _sum_tipo(cuadratura, "TAXI")
                cuadratura.jugados_dia  = _sum_tipo(cuadratura, "JUGADOS")
                cuadratura.otros_1_dia  = _sum_tipo(cuadratura, "OTROS")

                # Recalcular totales ahora que está todo seteado
                _recalcular_totales(cuadratura)

                cuadratura.actualizado_el = timezone.now()
                cuadratura.save()

                # ✅ ACTUALIZAR LA CUADRATURA DEL DÍA SIGUIENTE
                siguiente_dia = cuadratura.fecha + timedelta(days=1)
                cuadratura_siguiente = CuadraturaCajaDiaria.objects.filter(
                    sucursal=cuadratura.sucursal,
                    fecha=siguiente_dia
                ).first()

                if cuadratura_siguiente:
                    cuadratura_siguiente.caja = cuadratura.desglose_efectivo_total
                    cuadratura_siguiente.save()

            messages.success(request, "Cuadratura guardada (creada o actualizada) exitosamente.")
            return redirect("control:cuadratura_diaria_list")

        # Si el formulario no es válido
        sucursal_id = request.POST.get("sucursal")
        fecha_post = request.POST.get("fecha") or timezone.now().date()
        try:
            sucursal_obj = Sucursal.objects.get(pk=sucursal_id)
            caja_anterior, prestamos_acum_ant = _caja_anterior_en_ciclo(sucursal_obj, fecha_post if hasattr(fecha_post, 'year') else timezone.datetime.fromisoformat(str(fecha_post)).date())
            caja_inicial = int(sucursal_obj.caja_inicial or 0)
        except (Sucursal.DoesNotExist, ValueError, TypeError):
            sucursal_obj = None
            caja_anterior = 0
            caja_inicial = 0
            prestamos_acum_ant = 0

        messages.error(request, "Formulario inválido. Revise los datos enviados.")

    else:
        form = CuadraturaCajaDiariaForm(initial={"fecha": timezone.now().date()})
        sucursales = Sucursal.objects.filter(is_active=True).order_by("nombre")
        primera_sucursal = sucursales.first()

        if primera_sucursal:
            fecha = timezone.now().date()
            caja_anterior, prestamos_acum_ant = _caja_anterior_en_ciclo(primera_sucursal, fecha)
            caja_inicial = int(primera_sucursal.caja_inicial or 0)
        else:
            caja_anterior = 0
            caja_inicial = 0
            prestamos_acum_ant = 0

    # Verificar si todas las zonas del turno activo tienen líneas en el control
    fecha_hoy = timezone.localdate()
    zonas_faltantes = []
    turno_activo_caja = None
    for suc in Sucursal.objects.filter(is_active=True):
        turno_hoy = Turno.objects.filter(sucursal=suc, fecha=fecha_hoy, estado="Abierto").first()
        if turno_hoy:
            turno_activo_caja = turno_hoy
            control_hoy = ControlLecturas.objects.filter(sucursal=suc, fecha_trabajo=fecha_hoy).first()
            zonas_con_lineas = set()
            if control_hoy:
                zonas_con_lineas = set(
                    control_hoy.lineas.values_list("zona_id", flat=True).distinct()
                )
            for az in AsignacionTurnoZona.objects.filter(turno=turno_hoy).select_related("zona"):
                if az.zona_id not in zonas_con_lineas:
                    zonas_faltantes.append(az.zona.nombre)

    return render(request, "cuadratura_diaria/create.html", {
        "form": form,
        "sucursales": Sucursal.objects.filter(is_active=True).order_by("nombre"),
        "mes_default": timezone.localdate().strftime("%Y-%m"),
        "caja_anterior": caja_anterior,
        "caja_inicial": caja_inicial,
        "prestamos_acum_ant": prestamos_acum_ant,
        "zonas_faltantes": zonas_faltantes,
        "control_completo": len(zonas_faltantes) == 0,
    })
@login_required
@role_required(*ROLES_CUADRATURA)
def cuadratura_diaria_list(request):
    cuadraturas = CuadraturaCajaDiaria.objects.all().select_related("sucursal", "usuario")

    # -------------------------
    # Filtros (GET)
    # -------------------------
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    sucursal_id = request.GET.get("sucursal")

    if sucursal_id:
        cuadraturas = cuadraturas.filter(sucursal_id=sucursal_id)

    if fecha_desde:
        cuadraturas = cuadraturas.filter(fecha__gte=fecha_desde)

    if fecha_hasta:
        cuadraturas = cuadraturas.filter(fecha__lte=fecha_hasta)

    cuadraturas = cuadraturas.order_by("-fecha", "-creado_el")

    from django.core.paginator import Paginator
    paginator = Paginator(cuadraturas, 15)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(
        request,
        "cuadratura_diaria/list.html",
        {
            "cuadraturas": page_obj,
            "page_obj": page_obj,
            "sucursales": Sucursal.objects.filter(is_active=True).order_by("nombre"),
        },
    )



@login_required
def cuadratura_diaria_detail(request, pk):
    cuadratura = get_object_or_404(CuadraturaCajaDiaria, pk=pk)
    caja_anterior, _ = _caja_anterior_en_ciclo(cuadratura.sucursal, cuadratura.fecha)

    # Hora de apertura del turno: usa el turno vinculado o busca por sucursal+fecha
    turno_ref = cuadratura.turno or (
        Turno.objects.filter(sucursal=cuadratura.sucursal, fecha=cuadratura.fecha)
        .order_by("created_at")
        .first()
    )
    hora_apertura = turno_ref.created_at if turno_ref else None

    return render(request, "cuadratura_diaria/detail.html", {
        "cuadratura":    cuadratura,
        "caja_anterior": caja_anterior,
        "hora_apertura": hora_apertura,
    })
@login_required
def cuadratura_diaria_edit(request, pk):
    cuadratura = get_object_or_404(CuadraturaCajaDiaria, pk=pk)

    if request.method == "POST":
        form = CuadraturaCajaDiariaForm(request.POST, instance=cuadratura)
        if form.is_valid():
            c = form.save(commit=False)

            # Recalcular acumulados y totales igual que en create
            c.sorteos_acum = (c.sorteos_ant or 0) + (c.sorteos_dia or 0)
            c.gastos_acum = (c.gastos_ant or 0) + (c.gastos_dia or 0)
            c.sueldo_b_acum = (c.sueldo_b_ant or 0) + (c.sueldo_b_dia or 0)
            c.redbank_acum = (c.redbank_ant or 0) + (c.redbank_dia or 0)
            c.regalos_acum = (c.regalos_ant or 0) + (c.regalos_dia or 0)
            c.taxi_acum = (c.taxi_ant or 0) + (c.taxi_dia or 0)
            c.jugados_acum = (c.jugados_ant or 0) + (c.jugados_dia or 0)
            c.transfer_acum = (c.transfer_ant or 0) + (c.transfer_dia or 0)
            c.otros_1_acum = (c.otros_1_ant or 0) + (c.otros_1_dia or 0)
            c.otros_2_acum = (c.otros_2_ant or 0) + (c.otros_2_dia or 0)
            c.otros_3_acum = (c.otros_3_ant or 0) + (c.otros_3_dia or 0)
            c.descuadre_acum = (c.descuadre_ant or 0) + (c.descuadre_dia or 0)

            # Guardar campos de desglose de billetes
            for field in ['ef_20000', 'ef_10000', 'ef_5000', 'ef_2000', 'ef_1000', 'ef_monedas', 'ef_billetes_malos']:
                setattr(c, field, form.cleaned_data.get(field, 0))

            # Recalcular ganancia, numeral_dia, descuadre y total_efectivo correctamente
            _recalcular_totales(c)

            c.actualizado_el = timezone.now()
            c.save()

            # Propagar cambios en cascada a todos los días siguientes de la misma sucursal
            # Para que la caja_anterior de cada día siguiente quede actualizada
            dias_siguientes = CuadraturaCajaDiaria.objects.filter(
                sucursal=c.sucursal,
                fecha__gt=c.fecha
            ).order_by("fecha", "creado_el")

            for siguiente in dias_siguientes:
                _recalcular_totales(siguiente)
                siguiente.actualizado_el = timezone.now()
                siguiente.save()

            messages.success(request, "Cuadratura actualizada y días siguientes recalculados.")
            return redirect("control:cuadratura_diaria_detail", pk=c.pk)
        else:
            messages.error(request, "Formulario inválido.")
    else:
        form = CuadraturaCajaDiariaForm(instance=cuadratura)

    # Última caja guardada antes de esta (mismo día más antigua, o días anteriores)
    caja_anterior, prestamos_acum_ant = _caja_anterior_en_ciclo(cuadratura.sucursal, cuadratura.fecha)
    sucursales = Sucursal.objects.filter(is_active=True).order_by("nombre")

    return render(request, "cuadratura_diaria/create.html", {
        "form": form,
        "cuadratura": cuadratura,
        "caja_anterior": caja_anterior,
        "caja_inicial": int(cuadratura.sucursal.caja_inicial or 0),
        "prestamos_acum_ant": prestamos_acum_ant or 0,
        "mes_default": timezone.localdate().strftime("%Y-%m"),
        "sucursales": sucursales,
        "editar": True,
    })



@login_required
def cuadratura_diaria_recalcular_todo(request):
    """
    Recalcula total_efectivo y ganancia de TODAS las cuadraturas en orden cronológico.
    Esto propaga correctamente la cadena: caja_anterior → ganancia → total_efectivo → siguiente día.
    Solo accesible para admins.
    """
    if request.user.role != "admin":
        messages.error(request, "Solo los administradores pueden recalcular.")
        return redirect("control:cuadratura_diaria_list")

    if request.method != "POST":
        total = CuadraturaCajaDiaria.objects.count()
        return render(request, "cuadratura_diaria/recalcular.html", {"total": total})

    # Procesar en orden cronológico para que cada día vea el total_efectivo correcto del día anterior
    cuadraturas = CuadraturaCajaDiaria.objects.select_related("sucursal").order_by("fecha", "creado_el")
    actualizadas = 0

    for c in cuadraturas:
        try:
            _recalcular_totales(c)
            c.actualizado_el = timezone.now()
            c.save()
            actualizadas += 1
        except Exception as e:
            messages.warning(request, f"Error en {c.sucursal} {c.fecha}: {e}")

    messages.success(request, f"✅ {actualizadas} cuadraturas recalculadas correctamente.")
    return redirect("control:cuadratura_diaria_list")

@login_required
def cuadratura_diaria_delete(request, pk):
    cuadratura = get_object_or_404(CuadraturaCajaDiaria, pk=pk)

    if request.method == "POST":
        cuadratura.delete()
        messages.success(request, "Cuadratura eliminada.")
        return redirect("control:cuadratura_diaria_list")

    return render(request, "cuadratura_diaria/delete.html", {"cuadratura": cuadratura})



@login_required
def cuadratura_diaria_recalcular_todo(request):
    """
    Recalcula total_efectivo y ganancia de TODAS las cuadraturas en orden cronológico.
    Esto propaga correctamente la cadena: caja_anterior → ganancia → total_efectivo → siguiente día.
    Solo accesible para admins.
    """
    if request.user.role != "admin":
        messages.error(request, "Solo los administradores pueden recalcular.")
        return redirect("control:cuadratura_diaria_list")

    if request.method != "POST":
        total = CuadraturaCajaDiaria.objects.count()
        return render(request, "cuadratura_diaria/recalcular.html", {"total": total})

    # Procesar en orden cronológico para que cada día vea el total_efectivo correcto del día anterior
    cuadraturas = CuadraturaCajaDiaria.objects.select_related("sucursal").order_by("fecha", "creado_el")
    actualizadas = 0

    for c in cuadraturas:
        try:
            _recalcular_totales(c)
            c.actualizado_el = timezone.now()
            c.save()
            actualizadas += 1
        except Exception as e:
            messages.warning(request, f"Error en {c.sucursal} {c.fecha}: {e}")

    messages.success(request, f"✅ {actualizadas} cuadraturas recalculadas correctamente.")
    return redirect("control:cuadratura_diaria_list")

@login_required
def cuadratura_diaria_delete(request, pk):
    cuadratura = get_object_or_404(CuadraturaCajaDiaria, pk=pk)

    if request.method == "POST":
        cuadratura.delete()
        messages.success(request, "Cuadratura eliminada.")
        return redirect("control:cuadratura_diaria_list")

    return render(request, "cuadratura_diaria/delete.html", {"cuadratura": cuadratura})

@login_required
def cuadratura_diaria_export_excel(request):
    qs = CuadraturaCajaDiaria.objects.all().select_related("sucursal", "usuario")

    # mismos filtros que el listado
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    sucursal_id = request.GET.get("sucursal")

    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    qs = qs.order_by("-fecha", "-creado_el")

    wb = Workbook()
    ws = wb.active
    ws.title = "Cuadraturas Caja"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Fecha", "Sucursal", "Usuario",
        "Zona 1", "Zona 2", "Numeral Total",
        "Gastos Día", "Ganancia",
        "Caja", "Retiro Diario", "Total Efectivo",
        "Observaciones",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for c in qs:
        ws.cell(row=row, column=1).value = c.fecha.strftime("%d/%m/%Y") if c.fecha else ""
        ws.cell(row=row, column=2).value = c.sucursal.nombre if c.sucursal else ""
        ws.cell(row=row, column=3).value = getattr(c.usuario, "nombre", "") if c.usuario else ""

        # Estas referencias no existen en models -> CuadraturaCajaDiaria
       # ws.cell(row=row, column=4).value = int(c.zona_1 or 0)
       # ws.cell(row=row, column=5).value = int(c.zona_2 or 0)
       # ws.cell(row=row, column=6).value = int(c.numeral_total or 0)

       # Creo que estas son las que corresponden
        ws.cell(row=row, column=4).value = int(c.numeral_dia or 0)
        ws.cell(row=row, column=5).value = int(c.numeral_acumulado or 0)
        # No se es total_efectivo falta por correguir
        ws.cell(row=row, column=6).value = int(c.total_efectivo or 0)
        
        
        # método del modelo
        ws.cell(row=row, column=7).value = int(c.total_gastos_dia() or 0)

        ws.cell(row=row, column=8).value = int(c.ganancia or 0)

        ws.cell(row=row, column=9).value = int(c.caja or 0)
        ws.cell(row=row, column=10).value = int(c.retiro_diario or 0)
        ws.cell(row=row, column=11).value = int(c.total_efectivo or 0)

        ws.cell(row=row, column=12).value = (c.observaciones or "")
        row += 1

    # ancho automático
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    filename = f"cuadraturas_caja_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

# ============================================
# ENCUADRE CAJA (SOLO ADMIN)
# ============================================
from django.db.models import Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment



@login_required
@user_passes_test(is_admin, login_url="/login/")
def encuadre_create(request):
    if request.method == "POST":
        form = EncuadreCajaAdminForm(request.POST)
        if form.is_valid():
            encuadre = form.save(commit=False)

            encuadre.usuario_admin = request.user
            encuadre.sucursal = form.cleaned_data["sucursal"]


            # total caja
            encuadre.total_caja = (
                (encuadre.caja_numeral or 0)
                + (encuadre.prestamos or 0)
                - (encuadre.redbank_retiros or 0)
            )

            # total entregado
            total_billetes = (
                (encuadre.billete_20000 or 0) * 20000
                + (encuadre.billete_10000 or 0) * 10000
                + (encuadre.billete_5000 or 0) * 5000
                + (encuadre.billete_2000 or 0) * 2000
                + (encuadre.billete_1000 or 0) * 1000
            )
            total_entregado = total_billetes + (encuadre.monedas_total or 0)

            # descuadre
            encuadre.descuadre = encuadre.total_caja - total_entregado

            encuadre.save()
            messages.success(request, "Encuadre de caja creado exitosamente.")
            return redirect("control:encuadre_list")
        else:
            messages.error(request, "Formulario inválido. Revise los datos enviados.")
    else:
        form = EncuadreCajaAdminForm(initial={"fecha": timezone.now().date()})

    return render(request, "encuadre/create.html", {"form": form})


@login_required
@user_passes_test(is_admin, login_url="/login/")
def encuadre_list(request):
    encuadres = EncuadreCajaAdmin.objects.all().select_related("sucursal", "usuario_admin")
    from django.core.paginator import Paginator
    paginator = Paginator(encuadres, 15)
    encuadres = paginator.get_page(request.GET.get("page", 1))
    return render(request, "encuadre/list.html", {"encuadres": encuadres, "page_obj": encuadres})


@login_required
@user_passes_test(is_admin, login_url="/login/")
def encuadre_detail(request, pk):
    encuadre = get_object_or_404(EncuadreCajaAdmin, pk=pk)
    return render(request, "encuadre/detail.html", {"encuadre": encuadre})


# ============================================
# DASHBOARD (SOLO ADMIN)
# ============================================

@login_required
@role_required(*ROLES_DASHBOARD)
def dashboard_view(request):
    from decimal import Decimal
    hoy = timezone.localdate()

    # ── Totales globales ────────────────────────────────────────────────────
    todas_lecturas = LecturaMaquina.objects.all()
    global_kpis = todas_lecturas.aggregate(
        total_entrada = Sum("entrada_dia"),
        total_salida  = Sum("salida_dia"),
        total_neto    = Sum("total"),
    )
    for k in global_kpis:
        global_kpis[k] = global_kpis[k] or 0

    if global_kpis["total_entrada"] and global_kpis["total_entrada"] > 0:
        global_kpis["rtp_global"] = round(
            global_kpis["total_salida"] / global_kpis["total_entrada"] * 100, 2
        )
    else:
        global_kpis["rtp_global"] = 0

    # ── Retiros y prestamos acumulados ──────────────────────────────────────
    total_retiros   = AsignacionTurnoZona.objects.aggregate(s=Sum("retiros"))["s"] or 0
    total_prestamos = AsignacionTurnoZona.objects.aggregate(s=Sum("prestamo"))["s"] or 0

    # ── Top/Bottom RTP de locales (último turno + histórico) ────────────────
    locales_rtp = []
    for suc in Sucursal.objects.filter(is_active=True):
        ultimo_turno = (
            Turno.objects.filter(sucursal=suc)
            .order_by("-fecha", "-created_at")
            .first()
        )
        # RTP último turno
        if ultimo_turno:
            agg = LecturaMaquina.objects.filter(turno=ultimo_turno).aggregate(
                ent=Sum("entrada_dia"), sal=Sum("salida_dia")
            )
            t_ent = agg["ent"] or 0
            t_sal = agg["sal"] or 0
            rtp_turno = round(t_sal / t_ent * 100, 2) if t_ent > 0 else None
        else:
            rtp_turno = None

        # RTP histórico
        agg_h = LecturaMaquina.objects.filter(sucursal=suc).aggregate(
            ent=Sum("entrada_dia"), sal=Sum("salida_dia")
        )
        h_ent = agg_h["ent"] or 0
        h_sal = agg_h["sal"] or 0
        rtp_hist = round(h_sal / h_ent * 100, 2) if h_ent > 0 else None

        locales_rtp.append({
            "sucursal":     suc,
            "ultimo_turno": ultimo_turno,
            "rtp_turno":    rtp_turno,
            "rtp_hist":     rtp_hist,
        })

    # Separar los que tienen datos para ordenar
    turno_con_datos = [x for x in locales_rtp if x["rtp_turno"] is not None]
    hist_con_datos  = [x for x in locales_rtp if x["rtp_hist"]  is not None]

    rtp_por_local = {
        "turno_top":    sorted(turno_con_datos, key=lambda x: x["rtp_turno"], reverse=True),
        "turno_bottom": sorted(turno_con_datos, key=lambda x: x["rtp_turno"]),
        "hist_top":     sorted(hist_con_datos,  key=lambda x: x["rtp_hist"],  reverse=True),
        "hist_bottom":  sorted(hist_con_datos,  key=lambda x: x["rtp_hist"]),
    }

    # ── Máquinas malas (Mantenimiento) ─────────────────────────────────────
    maquinas_malas = list(
        Maquina.objects.filter(estado__in=["Mantenimiento", "Retirada"], is_active=True)
        .select_related("sucursal", "zona")
        .annotate(estado_ord=_estado_ord())
        .order_by("estado_ord", "sucursal__nombre", "numero_maquina")
    )
    total_maquinas_malas = len(maquinas_malas)

    # ── Redbank total (acumulado histórico de cierres) ──────────────────────
    total_redbank = CierreTurno.objects.aggregate(s=Sum("redbank_retiros"))["s"] or 0

    # ── Desglose por sucursal para KPIs desplegables ────────────────────────
    detalles_kpi = []
    for suc in Sucursal.objects.filter(is_active=True).order_by("nombre"):
        ret_suc = AsignacionTurnoZona.objects.filter(
            turno__sucursal=suc
        ).aggregate(s=Sum("retiros"))["s"] or 0
        prest_suc = AsignacionTurnoZona.objects.filter(
            turno__sucursal=suc
        ).aggregate(s=Sum("prestamo"))["s"] or 0
        rb_suc = CierreTurno.objects.filter(
            sucursal=suc
        ).aggregate(s=Sum("redbank_retiros"))["s"] or 0
        # RTP histórico ya está en locales_rtp
        rtp_entry = next((x for x in locales_rtp if x["sucursal"].pk == suc.pk), None)
        detalles_kpi.append({
            "sucursal":  suc,
            "retiros":   ret_suc,
            "prestamos": prest_suc,
            "redbank":   rb_suc,
            "rtp_turno": rtp_entry["rtp_turno"] if rtp_entry else None,
            "rtp_hist":  rtp_entry["rtp_hist"]  if rtp_entry else None,
        })

    # ── Data por sucursal ───────────────────────────────────────────────────
    sucursales_data = []
    for sucursal in Sucursal.objects.filter(is_active=True).order_by("nombre"):
        turno_activo = Turno.objects.filter(
            sucursal=sucursal, estado="Abierto"
        ).select_related("usuario").first()

        # Buscar la primer sesión del día del turno activo en esa sucursal
        # (el turno se abre después del login, así que no filtramos por turno)
        if turno_activo:
            # Primero intentar con sucursal asignada
            ultima_sesion = RegistroSesion.objects.filter(
                sucursal=sucursal,
                fecha=turno_activo.fecha,
            ).order_by("hora_inicio").first()
            # Si no hay, buscar por usuario del turno en esa fecha (sucursal puede ser null al login)
            if not ultima_sesion and turno_activo.usuario:
                ultima_sesion = RegistroSesion.objects.filter(
                    usuario=turno_activo.usuario,
                    fecha=turno_activo.fecha,
                ).order_by("hora_inicio").first()
        else:
            ultima_sesion = RegistroSesion.objects.filter(
                sucursal=sucursal
            ).order_by("-hora_inicio").first()

        lecs_turno = LecturaMaquina.objects.filter(
            sucursal=sucursal, turno=turno_activo
        ) if turno_activo else LecturaMaquina.objects.none()

        turno_ent = lecs_turno.aggregate(s=Sum("entrada_dia"))["s"] or 0
        turno_sal = lecs_turno.aggregate(s=Sum("salida_dia"))["s"] or 0
        turno_tot = lecs_turno.aggregate(s=Sum("total"))["s"] or 0
        rtp_dia   = round(turno_sal / turno_ent * 100, 2) if turno_ent > 0 else 0

        # ── Stats del último turno para el botón (activo o último cerrado) ──
        if turno_activo:
            btn_numeral  = turno_ent
            btn_ganancia = turno_tot
            btn_rtp      = rtp_dia
        else:
            ultimo_turno_cerrado = (
                Turno.objects.filter(sucursal=sucursal, estado="Cerrado")
                .order_by("-fecha", "-created_at")
                .first()
            )
            if ultimo_turno_cerrado:
                _lec = LecturaMaquina.objects.filter(turno=ultimo_turno_cerrado).aggregate(
                    ent=Sum("entrada_dia"), sal=Sum("salida_dia"), tot=Sum("total")
                )
                _ent = _lec["ent"] or 0
                _sal = _lec["sal"] or 0
                btn_numeral  = _ent
                btn_ganancia = _lec["tot"] or 0
                btn_rtp      = round(_sal / _ent * 100, 2) if _ent > 0 else 0
            else:
                btn_numeral = btn_ganancia = btn_rtp = None

        acum = LecturaMaquina.objects.filter(sucursal=sucursal).aggregate(
            ent=Sum("entrada_dia"), sal=Sum("salida_dia"), tot=Sum("total")
        )
        acum_ent = acum["ent"] or 0
        acum_sal = acum["sal"] or 0
        acum_tot = acum["tot"] or 0
        rtp_acum = round(acum_sal / acum_ent * 100, 2) if acum_ent > 0 else 0

        asig_qs = AsignacionTurnoZona.objects.filter(turno=turno_activo) if turno_activo else AsignacionTurnoZona.objects.none()
        retiros_turno   = asig_qs.aggregate(s=Sum("retiros"))["s"] or 0
        prestamos_turno = asig_qs.aggregate(s=Sum("prestamo"))["s"] or 0

        maquinas_sucursal = list(
            Maquina.objects.filter(sucursal=sucursal, is_active=True)
            .select_related("zona")
            .annotate(estado_ord=_estado_ord())
            .order_by("estado_ord", "zona__orden", "numero_maquina")
        )

        ultimo_control = ControlLecturas.objects.filter(sucursal=sucursal).first()

        sucursales_data.append({
            "sucursal":        sucursal,
            "turno_activo":    turno_activo,
            "ultima_sesion":   ultima_sesion,
            "turno_ent":       turno_ent,
            "turno_sal":       turno_sal,
            "turno_tot":       turno_tot,
            "rtp_dia":         rtp_dia,
            "acum_ent":        acum_ent,
            "acum_sal":        acum_sal,
            "acum_tot":        acum_tot,
            "rtp_acum":        rtp_acum,
            "retiros_turno":   retiros_turno,
            "prestamos_turno": prestamos_turno,
            "maquinas":        maquinas_sucursal,
            "ultimo_control":  ultimo_control,
            "btn_numeral":     btn_numeral,
            "btn_ganancia":    btn_ganancia,
            "btn_rtp":         btn_rtp,
        })

    return render(request, "dashboard.html", {
        "hoy":                  hoy,
        "global_kpis":          global_kpis,
        "total_retiros":        total_retiros,
        "total_prestamos":      total_prestamos,
        "rtp_por_local":        rtp_por_local,
        "sucursales_data":      sucursales_data,
        "maquinas_malas":       maquinas_malas,
        "total_maquinas_malas": total_maquinas_malas,
        "total_redbank":        total_redbank,
        "detalles_kpi":         detalles_kpi,
    })


# ============================================
# TURNO
# ============================================

@login_required
@role_required(*ROLES_TURNO)
def turno_view(request):
    turno_abierto = Turno.objects.filter(usuario=request.user, estado="Abierto").first()

    if request.method == "POST":
        if not turno_abierto:
            form = TurnoForm(request.POST, user=request.user)
            if form.is_valid():
                turno = form.save(commit=False)
                turno.usuario = request.user
                turno.estado = "Abierto"
                if request.user.role == "usuario" and request.user.sucursal:
                    turno.sucursal = request.user.sucursal
                try:
                    turno.full_clean()
                    turno.save()
                    messages.success(request, "Turno iniciado exitosamente.")
                    return redirect("control:turno")
                except Exception as e:
                    messages.error(request, str(e))
        else:
            messages.warning(request, "Ya tiene un turno abierto.")
    else:
        # Para admin: form con fecha de hoy pre-cargada pero editable
        # Para usuario: form normal (sucursal bloqueada si tiene asignada)
        initial = {}
        if request.user.role == "admin":
            from datetime import date
            initial["fecha"] = date.today()
        form = TurnoForm(user=request.user, initial=initial) if not turno_abierto else None

    cantidad_lecturas = 0
    zonas = []
    asig_zonas = {}
    asig_slots = {}
    usuarios_sucursal = []
    cant_redbank = 0
    cant_servicios = 0

    if turno_abierto:
        cantidad_lecturas = LecturaMaquina.objects.filter(turno=turno_abierto).count()
        zonas = list(Zona.objects.filter(sucursal=turno_abierto.sucursal, is_active=True).order_by("orden", "nombre"))
        usuarios_sucursal = list(Usuario.objects.filter(is_active=True, role="asistente", sucursales = turno_abierto.sucursal).order_by("nombre"))

        for az in AsignacionTurnoZona.objects.filter(turno=turno_abierto).select_related("usuario", "zona"):
            asig_zonas[az.zona_id] = az

        slots = list(AsignacionTurnoSlot.objects.filter(turno=turno_abierto).select_related("usuario"))
        redbank_slots = [s for s in slots if s.tipo == "redbank"]
        servicio_slots = [s for s in slots if s.tipo == "servicio"]
        if redbank_slots:
            cant_redbank = max(s.numero for s in redbank_slots)
        if servicio_slots:
            cant_servicios = max(s.numero for s in servicio_slots)
        for s in slots:
            asig_slots[(s.tipo, s.numero)] = s

    context = {
        "turno_abierto": turno_abierto,
        "cantidad_lecturas": cantidad_lecturas,
        "form": form,
        "sucursales": Sucursal.objects.filter(is_active=True).order_by("nombre"),
        "zonas": zonas,
        "usuarios_sucursal": usuarios_sucursal,
        "asig_zonas": asig_zonas,
        "asig_slots": asig_slots,
        "cant_redbank": cant_redbank,
        "cant_servicios": cant_servicios,
        "cant_redbank_range": list(range(1, cant_redbank + 1)),
        "cant_servicios_range": list(range(1, cant_servicios + 1)),
    }

    if request.user.role == "asistente":
        return render(request, "turno_asistente.html", context)
    return render(request, "turno.html", context)


@login_required
@transaction.atomic
def guardar_asignaciones(request, turno_id):
    if request.method != "POST":
        return redirect("control:turno")

    turno = get_object_or_404(Turno, id=turno_id, estado="Abierto")
    if request.user.role != "admin" and turno.usuario_id != request.user.id:
        messages.error(request, "No autorizado.")
        return redirect("control:turno")

    post = request.POST

    # Zonas
    zonas = Zona.objects.filter(sucursal=turno.sucursal, is_active=True)
    # Pre-fetch existing assignments to check banano lock
    existing_az = {az.zona_id: az for az in AsignacionTurnoZona.objects.filter(turno=turno)}
    for zona in zonas:
        uid_str  = post.get("zona_%d_usuario" % zona.id, "")
        banano_nuevo = int(post.get("zona_%d_banano" % zona.id, 0) or 0)
        prestamo = int(post.get("zona_%d_prestamo" % zona.id, 0) or 0)
        retiros  = int(post.get("zona_%d_retiros"  % zona.id, 0) or 0)
        uid = int(uid_str) if uid_str.isdigit() else None
        # Banano bloqueado: si ya fue guardado con valor > 0, no se puede modificar
        az_prev = existing_az.get(zona.id)
        banano = az_prev.banano if (az_prev and az_prev.banano > 0) else banano_nuevo
        AsignacionTurnoZona.objects.update_or_create(
            turno=turno, zona=zona,
            defaults={"usuario_id": uid, "banano": banano, "prestamo": prestamo, "retiros": retiros}
        )

    # Redbank
    cant_redbank = int(post.get("cant_redbank", 0) or 0)
    AsignacionTurnoSlot.objects.filter(turno=turno, tipo="redbank", numero__gt=cant_redbank).delete()
    for i in range(1, cant_redbank + 1):
        uid_str = post.get("redbank_%d_usuario" % i, "")
        uid = int(uid_str) if uid_str.isdigit() else None
        AsignacionTurnoSlot.objects.update_or_create(
            turno=turno, tipo="redbank", numero=i,
            defaults={"usuario_id": uid}
        )

    # Servicios
    cant_servicios = int(post.get("cant_servicios", 0) or 0)
    AsignacionTurnoSlot.objects.filter(turno=turno, tipo="servicio", numero__gt=cant_servicios).delete()
    for i in range(1, cant_servicios + 1):
        uid_str = post.get("servicio_%d_usuario" % i, "")
        uid = int(uid_str) if uid_str.isdigit() else None
        AsignacionTurnoSlot.objects.update_or_create(
            turno=turno, tipo="servicio", numero=i,
            defaults={"usuario_id": uid}
        )

    messages.success(request, "Asignaciones guardadas correctamente.")

    # Verificar asistentes que quedaron sin asignación
    _cerrar_sesion_asistentes_sin_asignacion(turno)

    return redirect("control:turno")


def _cerrar_sesion_asistentes_sin_asignacion(turno):
    """
    Después de guardar asignaciones, revisa asistentes con turno abierto
    en esta sucursal que ya no tienen ninguna asignación (zona ni slot).
    Les cierra el turno, el RegistroSesion y elimina todas sus sesiones Django.
    """
    from django.utils import timezone as tz
    from django.contrib.sessions.models import Session
    from decimal import Decimal
    from django.db.models import Sum as _Sum

    turnos_asistentes = Turno.objects.filter(
        sucursal=turno.sucursal,
        estado="Abierto",
        usuario__role="asistente"
    ).select_related("usuario")

    for t in turnos_asistentes:
        asistente = t.usuario
        tiene_zona = AsignacionTurnoZona.objects.filter(turno=turno, usuario=asistente).exists()
        tiene_slot = AsignacionTurnoSlot.objects.filter(turno=turno, usuario=asistente).exists()

        if not tiene_zona and not tiene_slot:
            # Cerrar turno
            total = LecturaMaquina.objects.filter(turno=t).aggregate(s=_Sum("total"))["s"] or 0
            t.estado = "Cerrado"
            t.total_cierre = Decimal(str(total))
            t.save()

            # Cerrar RegistroSesion con motivo
            try:
                RegistroSesion.objects.filter(
                    usuario=asistente, hora_cierre__isnull=True
                ).update(hora_cierre=tz.now(), motivo_cierre="sin_asignacion")
            except Exception:
                pass

            # Eliminar TODAS las sesiones Django del asistente
            # y guardar username en cada sesión para que el login muestre la alerta
            keys = []
            for session in Session.objects.filter(expire_date__gte=tz.now()):
                try:
                    data = session.get_decoded()
                    if str(data.get("_auth_user_id")) == str(asistente.pk):
                        # Inyectar username en la sesión antes de borrarla
                        # (no aplica porque la sesión se borra, usamos ?u= en la URL de login)
                        keys.append(session.session_key)
                except Exception:
                    pass
            if keys:
                Session.objects.filter(session_key__in=keys).delete()
            # Guardar flag en caché — cuando el asistente llegue al login
            # Django lo redirige con ?next=... y nosotros añadimos ?uid= al siguiente redirect
            try:
                from django.core.cache import cache
                cache.set(f"forzado_sin_asig_{asistente.pk}", True, timeout=300)
            except Exception:
                pass


# ============================================
# CUADRATURA ZONA
# ============================================

DENOMINACIONES_ZONA = [
    {"key": "20000", "label": "20.000", "valor": 20000},
    {"key": "10000", "label": "10.000", "valor": 10000},
    {"key": "5000",  "label": "5.000",  "valor": 5000},
    {"key": "2000",  "label": "2.000",  "valor": 2000},
    {"key": "1000",  "label": "1.000",  "valor": 1000},
]


@login_required
def cuadratura_zona_view(request, turno_id):
    try:
        from .models import CuadraturaZona
    except ImportError:
        CuadraturaZona = None
    turno = get_object_or_404(Turno, id=turno_id)

    user = request.user

    # Asistente: buscar asignaciones en el turno del encargado de la misma sucursal
    if user.role == 'asistente':
        turno_encargado = Turno.objects.filter(
            sucursal=turno.sucursal, estado="Abierto", usuario__role="encargado"
        ).first()
        turno_ref = turno_encargado if turno_encargado else turno
        zona_ids = AsignacionTurnoZona.objects.filter(
            turno=turno_ref, usuario=user
        ).values_list('zona_id', flat=True)
        zonas = Zona.objects.filter(id__in=zona_ids, is_active=True).order_by("orden", "nombre")
    else:
        turno_ref = turno
        zonas = Zona.objects.filter(
            sucursal=turno.sucursal, is_active=True
        ).order_by("orden", "nombre")

    asig_zonas = {
        az.zona_id: az
        for az in AsignacionTurnoZona.objects.filter(turno=turno_ref).select_related("usuario", "zona")
    }

    cz_map = {}
    if CuadraturaZona is not None:
        cz_map = {
            cz.zona_id: cz
            for cz in CuadraturaZona.objects.filter(turno=turno_ref)
        }

    lecturas_raw = (
        LecturaMaquina.objects
        .filter(turno=turno_ref)
        .select_related("zona", "maquina")
        .annotate(estado_ord=Case(
            When(maquina__estado="Mantenimiento", then=Value(0)),
            When(maquina__estado="Retirada",      then=Value(1)),
            default=Value(2),
            output_field=IntegerField(),
        ))
        .order_by("zona__orden", "zona__nombre", "estado_ord", "numero_maquina")
    )
    lecturas_por_zona = {}
    for lec in lecturas_raw:
        lecturas_por_zona.setdefault(lec.zona_id, []).append(lec)

    zonas_data = []
    for zona in zonas:
        lecs = lecturas_por_zona.get(zona.id, [])
        total_lecs = sum(l.total for l in lecs)
        az = asig_zonas.get(zona.id)
        cz = cz_map.get(zona.id)
        zonas_data.append({
            "zona": zona,
            "asignacion": az,
            "cz": cz,
            "lecturas": lecs,
            "total_lecturas": total_lecs,
            "val_banano":   cz.banano   if cz else (az.banano   if az else 0),
            "val_prestamo": cz.prestamo if cz else (az.prestamo if az else 0),
            "val_retiros":  cz.retiros  if cz else (az.retiros  if az else 0),
        })

    # Asistente: solo sus slots del turno del encargado
    slots_qs = AsignacionTurnoSlot.objects.filter(turno=turno_ref).select_related("usuario").order_by("tipo", "numero")
    if user.role == 'asistente':
        slots_qs = slots_qs.filter(usuario=user)
    slots = list(slots_qs)
    redbank_data  = [{"numero": s.numero, "usuario": s.usuario} for s in slots if s.tipo == "redbank"]
    servicio_data = [{"numero": s.numero, "usuario": s.usuario} for s in slots if s.tipo == "servicio"]

    return render(request, "cuadratura_zona/cuadratura_zona.html", {
        "turno": turno,
        "zonas_data": zonas_data,
        "redbank_data": redbank_data,
        "servicio_data": servicio_data,
        "denominaciones": DENOMINACIONES_ZONA,
    })


@login_required
@transaction.atomic
def guardar_cuadratura_zona(request, turno_id, zona_id):
    try:
        from .models import CuadraturaZona
    except ImportError:
        CuadraturaZona = None
    if request.method != "POST":
        return redirect("control:cuadratura_zona", turno_id=turno_id)

    turno = get_object_or_404(Turno, id=turno_id)
    zona  = get_object_or_404(Zona, id=zona_id)
    post  = request.POST

    banano   = int(post.get("banano",   0) or 0)
    prestamo = int(post.get("prestamo", 0) or 0)
    retiros  = int(post.get("retiros",  0) or 0)

    ef_20000 = int(post.get("ef_20000", 0) or 0)
    ef_10000 = int(post.get("ef_10000", 0) or 0)
    ef_5000  = int(post.get("ef_5000",  0) or 0)
    ef_2000  = int(post.get("ef_2000",  0) or 0)
    ef_1000  = int(post.get("ef_1000",  0) or 0)
    monedas  = int(post.get("ef_monedas", 0) or 0)
    total_ef = ef_20000 + ef_10000 + ef_5000 + ef_2000 + ef_1000 + monedas

    notas    = post.get("notas", "").strip()

    # numeral_dia = suma de lecturas de maquinas de esa zona en ese turno
    numeral_dia = LecturaMaquina.objects.filter(turno=turno, zona=zona).aggregate(
        total=Sum("total")
    )["total"] or 0

    # descuadre = banano + numeral_dia + prestamos - retiros - efectivo
    descuadre = banano + numeral_dia + prestamo - retiros - total_ef

    if CuadraturaZona is not None:
        CuadraturaZona.objects.update_or_create(
            turno=turno, zona=zona,
            defaults={
                "banano":   banano,
                "prestamo": prestamo,
                "retiros":  retiros,
                "ef_20000":  ef_20000,
                "ef_10000":  ef_10000,
                "ef_5000":   ef_5000,
                "ef_2000":   ef_2000,
                "ef_1000":   ef_1000,
                "monedas_monto": monedas,
                "detalle_entregado_total": total_ef,
                "notas":     notas,
                "descuadre": descuadre,
            }
        )

    AsignacionTurnoZona.objects.filter(turno=turno, zona=zona).update(
        banano=banano, prestamo=prestamo, retiros=retiros
    )

    messages.success(request, f"Cuadratura de {zona.nombre} guardada.")
    return redirect("control:cuadratura_zona", turno_id=turno_id)


@login_required
def cerrar_turno(request, turno_id):
    turno = get_object_or_404(Turno, id=turno_id, usuario=request.user)

    if turno.estado == "Abierto":
        total = LecturaMaquina.objects.filter(turno=turno).aggregate(
            total_sum=Sum("total")
        )["total_sum"] or 0

        turno.estado = "Cerrado"
        turno.total_cierre = Decimal(str(total))
        turno.save()

        messages.success(
            request, f"Turno cerrado exitosamente. Total: ${turno.total_cierre:,.0f}"
        )
    else:
        messages.warning(request, "El turno ya está cerrado.")

    if request.user.role in ('encargado', 'asistente'):
        logout(request)
        return redirect("control:login")

    return redirect("control:cierre_turno", turno_id=turno.id)

@login_required
@role_required(*ROLES_REGISTRO)
@readonly_for(*ROLES_READONLY)
def registro_view(request):
    turno_abierto = Turno.objects.filter(usuario=request.user, estado__iexact="abierto").first()
    zona_guardada_id = request.session.get("ultima_zona")

    # Asistente: usar turno del encargado como referencia para asignaciones y zonas
    if request.user.role == 'asistente' and turno_abierto:
        turno_ref = Turno.objects.filter(
            sucursal=turno_abierto.sucursal, estado="Abierto", usuario__role="encargado"
        ).first() or turno_abierto
    else:
        turno_ref = turno_abierto

    if request.method == "POST":
        if not turno_abierto:
            messages.error(
                request,
                'Debe iniciar un turno en la pestaña "Turno" antes de registrar lecturas.',
            )
            return redirect("control:turno")

        # ── Flags de reporte ──────────────────────────────────────────────────────
        reportar_maquina   = request.POST.get("reportar_maquina") == "1"
        numeral_visible    = request.POST.get("numeral_visible") == "1"
        observacion_rep    = request.POST.get("observacion_reporte", "").strip()

        form = LecturaMaquinaForm(request.POST, turno=turno_ref, usuario=request.user)

        if form.is_valid():
            lectura = form.save(commit=False)
            lectura.turno = turno_abierto
            lectura.usuario = request.user
            lectura.sucursal = turno_abierto.sucursal
            lectura.zona = form.cleaned_data["zona"]

            if lectura.maquina:
                lectura.numero_maquina = lectura.maquina.numero_maquina
                lectura.nombre_juego = lectura.maquina.nombre_juego

            try:
                # ── Si es reporte sin numeral visible, usar el último contador registrado ──
                from .utils import get_referencia_anterior
                entrada_ant, salida_ant, fuente = get_referencia_anterior(
                    lectura.maquina, turno_abierto.fecha
                )
                entrada_ant = int(entrada_ant or 0)
                salida_ant  = int(salida_ant or 0)

                if reportar_maquina and not numeral_visible:
                    lectura.entrada = entrada_ant
                    lectura.salida  = salida_ant

                lectura.total = int(lectura.entrada or 0) - int(lectura.salida or 0)

                # ── Si es reporte, agregar observación a la nota ──
                if reportar_maquina and observacion_rep:
                    prefijo = f"[REPORTE MANTENIMIENTO] {observacion_rep}"
                    lectura.nota = f"{prefijo} | {lectura.nota}" if lectura.nota else prefijo

                lectura.full_clean()

                # ── Validación: los contadores nunca deben retroceder ──
                entrada_nueva = int(lectura.entrada or 0)
                salida_nueva  = int(lectura.salida or 0)

                errores = []
                if not (reportar_maquina and not numeral_visible):
                    if entrada_nueva < entrada_ant:
                        errores.append(
                            f"Entrada ingresada ({entrada_nueva:,}) es MENOR al valor anterior ({entrada_ant:,}). "
                            f"Verifique que los datos estén bien ingresados."
                        )
                    if salida_nueva < salida_ant:
                        errores.append(
                            f"Salida ingresada ({salida_nueva:,}) es MENOR al valor anterior ({salida_ant:,}). "
                            f"Verifique que los datos estén bien ingresados."
                        )

                if errores:
                    for err in errores:
                        messages.error(request, f"⚠️ Máquina {lectura.numero_maquina}: {err}")
                else:
                    lectura.save()

                    # ── Cambiar estado de la máquina a Mantenimiento ──
                    if reportar_maquina:
                        Maquina.objects.filter(pk=lectura.maquina.pk).update(estado="Mantenimiento")

                    request.session["ultima_zona"] = lectura.zona.id

                    # Calcular la siguiente máquina en orden dentro de la zona
                    maquinas_zona = list(
                        Maquina.objects.filter(
                            zona=lectura.zona,
                            estado="Operativa",
                            sucursal__is_active=True,
                        ).order_by("numero_maquina").values_list("id", flat=True)
                    )
                    if lectura.maquina and lectura.maquina.id in maquinas_zona:
                        idx = maquinas_zona.index(lectura.maquina.id)
                        siguiente_idx = idx + 1
                        if siguiente_idx < len(maquinas_zona):
                            request.session["siguiente_maquina"] = maquinas_zona[siguiente_idx]
                        else:
                            request.session.pop("siguiente_maquina", None)
                    else:
                        request.session.pop("siguiente_maquina", None)

                    if reportar_maquina:
                        messages.warning(request, f"Máquina {lectura.maquina.numero_maquina} reportada y puesta en Mantenimiento.")
                    else:
                        messages.success(request, f"Lectura registrada. Máquina {lectura.maquina.numero_maquina} guardada.")
                    return redirect("control:registro")
            except IntegrityError:
                messages.error(request, "Ya existe una lectura para esa máquina en este turno.")
            except ValidationError as e:
                messages.error(request, "; ".join(e.messages))
        else:
            messages.error(request, "Formulario inválido. Revise los datos enviados.")

    else:
        if turno_abierto:
            siguiente_maquina_id = request.session.pop("siguiente_maquina", None)
            initial = {}
            if zona_guardada_id:
                initial["zona"] = zona_guardada_id
            if siguiente_maquina_id:
                initial["maquina"] = siguiente_maquina_id
            # Verificar si el usuario tiene zona asignada (buscar en turno de referencia)
            zonas_asignadas = AsignacionTurnoZona.objects.filter(
                turno=turno_ref, usuario=request.user
            ).select_related("zona")
            tiene_asignacion = zonas_asignadas.exists()

            # Auto-preseleccionar zona si solo tiene una asignada
            if not initial.get("zona") and tiene_asignacion and zonas_asignadas.count() == 1:
                initial["zona"] = zonas_asignadas.first().zona_id

            form = LecturaMaquinaForm(turno=turno_ref, usuario=request.user, initial=initial if initial else None)
        else:
            form = None
        siguiente_maquina_id = None  # ya fue consumida

    # Info de asignación de zona para mostrar aviso en template
    zonas_asignadas_usuario = []
    if turno_abierto:
        zonas_asignadas_usuario = list(
            AsignacionTurnoZona.objects.filter(turno=turno_ref, usuario=request.user)
            .select_related("zona").values_list("zona__nombre", flat=True)
        )

    context = {
        "turno_abierto": turno_abierto,
        "form": form,
        "zona_guardada": zona_guardada_id,
        "zonas_asignadas_usuario": zonas_asignadas_usuario,
        "tiene_asignacion": bool(zonas_asignadas_usuario),
    }

    return render(request, "registro.html", context)


# ============================================
# AJAX
# ============================================
@login_required
def _get_caja_anterior(sucursal: Sucursal, fecha):
    caja_ant, _ = _caja_anterior_en_ciclo(sucursal, fecha)
    return caja_ant

def ajax_cuadratura_detalles(request):
    sucursal_id = request.GET.get("sucursal_id")
    fecha = request.GET.get("fecha")

    if not sucursal_id or not fecha:
        return JsonResponse({"ok": False, "error": "Faltan parámetros."})

    sucursal = Sucursal.objects.filter(id=sucursal_id).first()
    if not sucursal:
        return JsonResponse({"ok": False, "error": "Sucursal no existe."})

    cuadratura = (
        CuadraturaCajaDiaria.objects
        .filter(sucursal=sucursal, fecha=fecha)
        .order_by("-creado_el")
        .first()
    )

    if not cuadratura:
        return JsonResponse({"ok": True, "exists": False, "detalles": []})

    detalles = list(
        CuadraturaDetalle.objects
        .filter(cuadratura=cuadratura)
        .values("tipo", "nombre", "monto")
        .order_by("tipo", "creado_en")
    )

    return JsonResponse({
        "ok": True,
        "exists": True,
        "id": cuadratura.id,
        "detalles": detalles,
        # opcional: para depurar
        "gastos_dia": cuadratura.gastos_dia,
        "sueldo_b_dia": cuadratura.sueldo_b_dia,
    })

def _caja_anterior_en_ciclo(sucursal, fecha):
    """
    Retorna (caja_anterior, prestamos_acum_ant).
    - Si fecha es el inicio del ciclo o anterior: devuelve (caja_inicial, 0).
    - Si no: busca la última CuadraturaCajaDiaria dentro del ciclo antes de fecha.
    """
    inicio_ciclo = get_inicio_ciclo(sucursal)

    # Primer día del nuevo ciclo → resetear todo
    if inicio_ciclo and fecha <= inicio_ciclo:
        return int(sucursal.caja_inicial or 0), 0

    qs = CuadraturaCajaDiaria.objects.filter(
        sucursal=sucursal,
        fecha__lt=fecha,
    )
    if inicio_ciclo:
        qs = qs.filter(fecha__gte=inicio_ciclo)

    prev = qs.order_by("-fecha", "-creado_el").first()
    if prev:
        return int(prev.total_efectivo or 0), int(prev.prestamos_acum or 0)
    return int(sucursal.caja_inicial or 0), 0

@login_required
def ajax_cuadratura_diaria_numerales(request):
    sucursal_id = request.GET.get("sucursal_id")
    fecha_str = request.GET.get("fecha")

    if not sucursal_id or not fecha_str:
        return JsonResponse({"ok": False, "error": "missing_params"})

    try:
        fecha = timezone.datetime.fromisoformat(fecha_str).date()
    except Exception:
        return JsonResponse({"ok": False, "error": "bad_date"})

    sucursal = Sucursal.objects.filter(id=sucursal_id, is_active=True).first()
    if not sucursal:
        return JsonResponse({"ok": False, "error": "bad_sucursal"})

    # === Cálculo principal ===
    numeral_dia, numeral_acumulado = calcular_numerales_caja(sucursal, fecha)

    caja_anterior, prestamos_acum_ant = _caja_anterior_en_ciclo(sucursal, fecha)

    return JsonResponse({
        "ok": True,
        "numeral_dia": numeral_dia,
        "numeral_acumulado": numeral_acumulado,
        "caja_anterior": caja_anterior,
        "caja_inicial": int(sucursal.caja_inicial or 0),
        "prestamos_acum_ant": prestamos_acum_ant,
    })

@login_required
def zonas_por_sucursal(request):
    sucursal_id = request.GET.get("sucursal_id")

    zonas = (
        Zona.objects
        .filter(sucursal_id=sucursal_id, is_active=True)
        .order_by("nombre")
    )

    data = [{"id": z.id, "nombre": z.nombre} for z in zonas]
    return JsonResponse(data, safe=False)

@login_required
def get_zonas_ajax(request, sucursal_id):
    zonas = Zona.objects.filter(
        sucursal_id=sucursal_id,
        is_active=True
    ).values("id", "nombre")
    return JsonResponse(list(zonas), safe=False)


@login_required
def get_maquinas_ajax(request, zona_id):
    maquinas = Maquina.objects.filter(
        zona_id=zona_id,
        estado="Operativa",
        zona__is_active=True
    ).values("id", "numero_maquina", "nombre_juego")
    return JsonResponse(list(maquinas), safe=False)


@login_required
def get_referencia_maquina_ajax(request, pk):
    """Devuelve el último numeral (entrada/salida) registrado para una máquina."""
    from .utils import get_referencia_anterior
    from django.utils import timezone
    maquina = get_object_or_404(Maquina, pk=pk)
    fecha = timezone.localdate()
    entrada_ant, salida_ant, fuente = get_referencia_anterior(maquina, fecha)
    return JsonResponse({
        "entrada_anterior": int(entrada_ant or 0),
        "salida_anterior":  int(salida_ant or 0),
        "fuente": fuente,
    })


# ============================================
# OCR
# ============================================

@csrf_exempt
def _preprocesar_imagen(imagen):
    """
    Preprocesado para pantallas de máquinas slot con fondo azul oscuro y texto blanco.
    Estrategia: aislar solo los píxeles muy brillantes (texto blanco/amarillo)
    y convertirlos a negro sobre fondo blanco para Tesseract.
    """
    from PIL import ImageEnhance, ImageFilter

    variantes = []
    imagen = imagen.convert("RGB")
    w, h = imagen.size

    # Auto-rotar: pantallas de slot son siempre apaisadas.
    # Si la foto viene en retrato (h > w), rotar 90° para que Tesseract lea mejor.
    if h > w:
        imagen = imagen.rotate(-90, expand=True)
        w, h = imagen.size
        print(f"[OCR] Imagen rotada a landscape: {w}x{h}")

    # Escalar a 1600px en el lado más largo
    lado = max(w, h)
    if lado < 1600:
        f = 1600 / lado
        imagen = imagen.resize((int(w*f), int(h*f)), Image.LANCZOS)
    elif lado > 2400:
        f = 2400 / lado
        imagen = imagen.resize((int(w*f), int(h*f)), Image.LANCZOS)

    gray = imagen.convert("L")

    # ── Variante 1: umbral alto — solo texto muy brillante (blanco puro) ──
    # En pantallas slot: texto blanco > 200, fondo azul < 100
    # Resultado: texto negro sobre blanco
    v1 = gray.point(lambda x: 0 if x > 180 else 255)
    variantes.append(v1)

    # ── Variante 2: umbral medio — captura texto ligeramente menos brillante
    v2 = gray.point(lambda x: 0 if x > 140 else 255)
    variantes.append(v2)

    # ── Variante 3: autocontraste + umbral alto ────────────────────────────
    v3 = ImageOps.autocontrast(gray, cutoff=3)
    v3 = v3.point(lambda x: 0 if x > 170 else 255)
    variantes.append(v3)

    # ── Variante 4: sharpening + autocontraste + umbral medio ─────────────
    v4 = gray.filter(ImageFilter.SHARPEN)
    v4 = ImageOps.autocontrast(v4, cutoff=2)
    v4 = v4.point(lambda x: 0 if x > 150 else 255)
    variantes.append(v4)

    # ── Variante 5: extracción RGB — aísla texto blanco/amarillo ──────────
    # Las pantallas slot usan texto blanco o amarillo sobre fondo oscuro.
    # Suma R+G por canal: blanco=alto en los 3, amarillo=R+G altos, B bajo.
    try:
        img_arr = np.array(imagen)   # shape: (H, W, 3)
        r = img_arr[:, :, 0].astype(np.int16)
        g = img_arr[:, :, 1].astype(np.int16)
        b = img_arr[:, :, 2].astype(np.int16)
        # Pixel "brillante": R y G ambos altos (blanco o amarillo)
        bright = ((r + g) > 360) & (r > 150) & (g > 150)
        mask = (bright.astype(np.uint8)) * 255
        v5 = Image.fromarray(255 - mask)  # negro sobre blanco
        variantes.append(v5)
    except Exception:
        pass  # si numpy falla por alguna razón, omitir esta variante

    # ── Variante 6: doble nitidez + umbral bajo (texto tenue) ─────────────
    v6 = gray.filter(ImageFilter.SHARPEN).filter(ImageFilter.SHARPEN)
    v6 = v6.point(lambda x: 0 if x > 120 else 255)
    variantes.append(v6)

    return variantes


def _ocr_texto(variante, config):
    """
    Corre tesseract y devuelve texto plano + data estructurada.
    Usa image_to_string para texto (más fiable) e image_to_data para estructura.
    """
    texto = pytesseract.image_to_string(variante, lang="eng", config=config)
    data  = pytesseract.image_to_data(variante, lang="eng", config=config, output_type=Output.DICT)
    # Print en consola del servidor para diagnóstico
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
    if lineas:
        print(f"[OCR] cfg={config[-5:]} texto={lineas[:4]}")
    else:
        print(f"[OCR] cfg={config[-5:]} SIN TEXTO")
    return texto, data


def _extraer_valores_de_texto(texto_plano, filas):
    """
    Estrategia principal: busca ENTRADAS/SALIDAS en filas estructuradas.
    Estrategia secundaria: regex sobre texto plano.
    Estrategia terciaria: busca columnas numéricas (pantallas con formato tabla).
    Devuelve (entrada, salida, entrada_raw, salida_raw) o Nones.
    """
    entrada = salida = entrada_raw = salida_raw = None

    # ── Estrategia 1: fila contiene LABEL + número en la misma línea ──────
    # Esta es la más confiable — "ENTRADAS $197.000" en una sola línea
    for f in filas:
        txt_norm = normalizar_texto_label(f["texto"])
        if entrada is None and es_linea_entrada(txt_norm) and f["numeros"]:
            r, s_str = extraer_monto_desde_tokens(f["numeros"])
            v = normalizar_valor(s_str)
            # Solo aceptar si el valor tiene sentido (>= 1000)
            if v and v >= 1000:
                entrada_raw, entrada = r, v
        if salida is None and es_linea_salida(txt_norm) and f["numeros"]:
            r, s_str = extraer_monto_desde_tokens(f["numeros"])
            v = normalizar_valor(s_str)
            if v and v >= 1000:
                salida_raw, salida = r, v

    # ── Fallback: label en fila i, número en fila i+1 ──────────────────────
    # Solo si la estrategia 1 no encontró algo
    if salida is None:
        for i in range(len(filas) - 1):
            if es_linea_salida(normalizar_texto_label(filas[i]["texto"])):
                nums = filas[i + 1]["numeros"]
                if nums:
                    salida_raw, salida_str = extraer_monto_desde_tokens(nums)
                    v = normalizar_valor(salida_str)
                    if v and v >= 1000:
                        salida_raw, salida = salida_raw, v
                        break
    if entrada is None:
        for i in range(len(filas) - 1):
            if es_linea_entrada(normalizar_texto_label(filas[i]["texto"])):
                nums = filas[i + 1]["numeros"]
                if nums:
                    entrada_raw, entrada_str = extraer_monto_desde_tokens(nums)
                    v = normalizar_valor(entrada_str)
                    if v and v >= 1000:
                        entrada_raw, entrada = entrada_raw, v
                        break

    # ── Estrategia 2: regex sobre texto plano ─────────────────────────────
    # Permite ruido entre el label y el número: "ALD eee 100" → "SALIDAS ... 402100"
    # [^\n]{0,25} = hasta 25 chars de ruido antes del número
    patron_entrada = r'(?:ENTRAD|UNTRAD|NTRADA)[A-Za-z]{0,4}[^\n]{0,25}?(\d[\d\.\,]{3,})'
    patron_salida  = r'(?:SALID|SAUD|SAIDA|ALIDA|ALD)[A-Za-z]{0,4}[^\n]{0,25}?(\d[\d\.\,]{3,})'

    if entrada is None:
        m = re.search(patron_entrada, texto_plano, re.IGNORECASE)
        if m:
            entrada_raw = m.group(1)
            entrada = normalizar_valor(entrada_raw)

    if salida is None:
        m = re.search(patron_salida, texto_plano, re.IGNORECASE)
        if m:
            salida_raw = m.group(1)
            salida = normalizar_valor(salida_raw)

    # ── Estrategia 3: buscar por proximidad en líneas ─────────────────────
    if entrada is None or salida is None:
        lineas = texto_plano.splitlines()
        for linea in lineas:
            norm = normalizar_texto_label(linea)
            # Buscar números de 4+ dígitos con posible separador de miles
            nums = re.findall(r"\d[\d\.\,]{3,}", linea)
            if not nums:
                continue
            if entrada is None and es_linea_entrada(norm):
                for n in nums:
                    v = normalizar_valor(n)
                    if v and v >= 1000:
                        entrada_raw = n
                        entrada = v
                        break
            if salida is None and es_linea_salida(norm):
                for n in nums:
                    v = normalizar_valor(n)
                    if v and v >= 1000:
                        salida_raw = n
                        salida = v
                        break

    return entrada, salida, entrada_raw, salida_raw


def ocr_lectura(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    try:
        # ── 1) Cargar imagen ───────────────────────────────────────────────
        if "imagen" in request.FILES:
            print(f"[OCR] Imagen recibida via FILES: {request.FILES['imagen'].size} bytes")
            imagen = Image.open(request.FILES["imagen"])
        else:
            data = json.loads(request.body.decode("utf-8"))
            b64 = data.get("image")
            if not b64:
                return JsonResponse({"success": False, "error": "No se recibió imagen"}, status=400)
            if "," in b64:
                b64 = b64.split(",", 1)[1]
            imagen = Image.open(io.BytesIO(base64.b64decode(b64)))

        # ── 2) Generar variantes preprocesadas ────────────────────────────
        variantes = _preprocesar_imagen(imagen)

        print(f"[OCR] {len(variantes)} variantes generadas, tamaños: {[v.size for v in variantes]}")

        # psm 11 = texto disperso — ideal para pantallas slot
        # psm 6  = bloque de texto uniforme — fallback
        # oem 1  = LSTM puro, más preciso en fuentes de pantalla
        WLIST = "-c tessedit_char_whitelist=0123456789$.,ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz "
        configs = [
            f"--oem 3 --psm 11 {WLIST}",
            f"--oem 3 --psm 6  {WLIST}",
            f"--oem 1 --psm 11 {WLIST}",
            f"--oem 3 --psm 11",
        ]

        # Candidatos recolectados de todas las combinaciones variante×config
        entradas_cand = []   # lista de (valor_int, raw_str)
        salidas_cand  = []
        mejor_texto   = ""
        errores_ocr   = []

        # ── 3) Correr TODAS las combinaciones y votar ──────────────────────
        for idx_v, variante in enumerate(variantes):
            for cfg in configs:
                try:
                    texto_plano, data_ocr = _ocr_texto(variante, cfg)
                    filas = agrupar_lineas(data_ocr)

                    e, s, er, sr = _extraer_valores_de_texto(texto_plano, filas)

                    if len(texto_plano.strip()) > len(mejor_texto.strip()):
                        mejor_texto = texto_plano

                    if e is not None:
                        entradas_cand.append((e, er))
                    if s is not None:
                        salidas_cand.append((s, sr))

                except Exception as ex_ocr:
                    errores_ocr.append(f"v{idx_v+1}: {ex_ocr}")
                    continue

                # Salida temprana: alta confianza alcanzada (≥3 votos coincidentes)
                if entradas_cand and salidas_cand:
                    top_e = Counter(v for v, _ in entradas_cand).most_common(1)[0][1]
                    top_s = Counter(v for v, _ in salidas_cand).most_common(1)[0][1]
                    if top_e >= 3 and top_s >= 3:
                        break
            else:
                continue
            break  # salir del loop externo si el inner hizo break

        # ── 4) Votar: elegir el valor más frecuente entre candidatos ──────
        def _votar(cand):
            if not cand:
                return None, None, 0
            c = Counter(v for v, _ in cand)
            valor, votos = c.most_common(1)[0]
            raw = next(r for v, r in cand if v == valor)
            return valor, raw, votos

        entrada, entrada_raw, votos_e = _votar(entradas_cand)
        salida,  salida_raw,  votos_s = _votar(salidas_cand)

        votos_min = min(votos_e or 0, votos_s or 0)
        confianza = "alta" if votos_min >= 3 else "media" if votos_min >= 2 else "baja"

        print(f"[OCR] Resultado: entrada={entrada}({votos_e}v) salida={salida}({votos_s}v) confianza={confianza}")

        # ── 5) Validar resultados ─────────────────────────────────────────
        if entrada is None or salida is None:
            faltantes = []
            if entrada is None: faltantes.append("ENTRADAS")
            if salida  is None: faltantes.append("SALIDAS")
            lineas_debug = [l.strip() for l in mejor_texto.splitlines() if l.strip()][:5]
            debug_visible = " | ".join(lineas_debug) if lineas_debug else "(sin texto)"
            return JsonResponse({
                "success":     False,
                "error":       f"No se pudieron leer: {', '.join(faltantes)}.",
                "debug_texto": debug_visible,
                "debug_full":  mejor_texto,
                "errores_ocr": errores_ocr,
            }, status=400)

        total = entrada - salida

        return JsonResponse({
            "success":     True,
            "entrada":     entrada,
            "salida":      salida,
            "total":       total,
            "entrada_raw": entrada_raw,
            "salida_raw":  salida_raw,
            "confianza":   confianza,
            "votos":       {"entrada": votos_e, "salida": votos_s},
            "mensaje":     f"Lectura procesada ({confianza} confianza, {votos_e}/{votos_s} votos)",
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({"success": False, "error": str(e)}, status=500)


def ocr_debug(request):
    """
    Endpoint de diagnóstico: recibe una imagen y devuelve las 3 variantes
    preprocesadas como base64 para ver exactamente qué ve Tesseract.
    URL: /ocr/debug/  (solo accesible para admin)
    """
    if not request.user.is_authenticated or request.user.role != "admin":
        return JsonResponse({"error": "No autorizado"}, status=403)
    if request.method != "POST":
        return JsonResponse({"error": "POST only"}, status=405)

    try:
        if "imagen" in request.FILES:
            imagen = Image.open(request.FILES["imagen"])
        else:
            data = json.loads(request.body.decode("utf-8"))
            b64 = data.get("image", "")
            if "," in b64:
                b64 = b64.split(",", 1)[1]
            imagen = Image.open(io.BytesIO(base64.b64decode(b64)))

        variantes = _preprocesar_imagen(imagen)
        resultado = []
        configs = ["--oem 3 --psm 6", "--oem 3 --psm 4"]

        for i, v in enumerate(variantes):
            buf = io.BytesIO()
            v.save(buf, format="PNG")
            img_b64 = base64.b64encode(buf.getvalue()).decode()

            # Correr OCR en esta variante
            texto_v = ""
            for cfg in configs:
                try:
                    data_ocr = pytesseract.image_to_data(v, lang="eng", config=cfg, output_type=Output.DICT)
                    texto_v = " ".join(t for t in data_ocr["text"] if t.strip())
                    if texto_v:
                        break
                except Exception:
                    pass

            resultado.append({
                "variante": i + 1,
                "imagen_base64": f"data:image/png;base64,{img_b64}",
                "ocr_texto": texto_v,
                "size": f"{v.size[0]}x{v.size[1]}",
            })

        return JsonResponse({"variantes": resultado})
    except Exception as e:
        import traceback
        return JsonResponse({"error": str(e), "trace": traceback.format_exc()}, status=500)


def agrupar_lineas(ocr_data, min_conf=25):
    """
    Agrupa palabras en líneas. min_conf descarta palabras con confianza baja
    (ruido OCR: caracteres fantasma, artefactos de compresión).
    """
    filas = {}
    n = len(ocr_data["text"])

    for i in range(n):
        txt = ocr_data["text"][i]
        if not txt or not txt.strip():
            continue

        # Filtrar palabras de baja confianza
        try:
            if int(ocr_data["conf"][i]) < min_conf:
                continue
        except (ValueError, TypeError, IndexError, KeyError):
            pass  # si no hay conf, incluir igualmente

        key = (ocr_data["block_num"][i], ocr_data["par_num"][i], ocr_data["line_num"][i])
        top = ocr_data["top"][i]

        if key not in filas:
            filas[key] = {"texto": "", "numeros": [], "top": top}
        else:
            filas[key]["top"] = min(filas[key]["top"], top)

        filas[key]["texto"] += " " + txt

        if any(c.isdigit() for c in txt):
            filas[key]["numeros"].append(txt)

    return sorted(filas.values(), key=lambda f: f["top"])


def normalizar_texto_label(s):
    if not s:
        return ""
    s = s.upper()
    reemplazos = {
        "0": "O", "1": "I", "5": "S", "4": "A", "|": "I",
        "Á": "A", "É": "E", "Í": "I", "Ó": "O", "Ú": "U",
        "/": ""
    }
    for k, v in reemplazos.items():
        s = s.replace(k, v)
    return s


def es_linea_entrada(txt):
    patrones = ["ENTRADA", "ENTRADAS", "NTRADA", "TRADA", "ENTRAD"]
    return any(p in txt for p in patrones)


def es_linea_salida(txt):
    # Incluye variantes por error OCR: SAUDAS (L→U), SAIDAS, SALUDA, etc.
    patrones = [
        "SALIDA", "SALIDAS", "SLIDA", "SALID", "SALD", "SALDA",
        "ALIDAS", "IDAS", "SAUDAS", "SAUDA", "SAIDA", "SAIIDAS",
        "SAUIDAS", "SLIDAS", "SAIIDA",
    ]
    return any(p in txt for p in patrones)


def extraer_monto_desde_tokens(tokens):
    raw = " ".join(tokens)
    # Buscar el patrón de número chileno: $197.000 o 197.000 o 197000
    # Primero intentar con formato de puntos (separador de miles chileno)
    m = re.search(r'[\$S]?\s*(\d{1,3}(?:\.\d{3})+)', raw)
    if m:
        solo_digitos = re.sub(r"[^\d]", "", m.group(1))
        return m.group(1), solo_digitos

    # Si no hay puntos, buscar número de 4+ dígitos
    m2 = re.search(r'(\d{4,})', raw)
    if m2:
        return m2.group(1), m2.group(1)

    # Fallback: quitar todo lo que no sea dígito
    solo_digitos = re.sub(r"[^\d]", "", raw)
    if len(solo_digitos) >= 4:
        return raw, solo_digitos

    candidato = seleccionar_numero_correcto(tokens)
    return candidato, candidato


def seleccionar_numero_correcto(nums):
    nums_limpios = [(n, len(re.sub(r"[^\d]", "", n))) for n in nums]
    nums_limpios.sort(key=lambda x: x[1], reverse=True)
    return nums_limpios[0][0] if nums_limpios else ""


def normalizar_valor(raw):
    if raw is None:
        return None
    s = str(raw).strip()

    # Limpiar símbolo peso ($, S, y dígitos que son $ mal leído: 5, 3, 6, 2)
    s = re.sub(r"^[\$sS]\s*", "", s).strip()
    # A veces el OCR lee $ como un dígito (5, 3, 6, 2) — detectarlo por contexto:
    # Si el número empieza con esos dígitos Y tiene 7+ dígitos → el primero es falso
    if re.match(r"^[23456]\d{6,}$", s):
        s = s[1:]

    # Formato chileno: 197.000 o 197,000 → separador de miles
    # Detectar patrón NNN.NNN o NNN,NNN (grupos exactos de 3)
    if re.match(r"^\d{1,3}([.,]\d{3})+$", s):
        s = re.sub(r"[.,]", "", s)
    else:
        s = re.sub(r"[^\d]", "", s)

    if not s:
        return None

    # Corregir "5" inicial = "$" mal leído por OCR
    if s.startswith("5") and len(s) >= 6 and not s.startswith("50"):
        s = s[1:]

    # Corrección de último dígito erróneo: si termina en múltiplo de 10
    # con variación de ±6 (ej: 197006 → probablemente 197000)
    # Solo aplica si el número tiene 6+ dígitos
    try:
        v = int(s)
        if len(s) >= 6:
            resto = v % 10
            if resto in (1, 2, 6, 7, 8, 9):
                # Último dígito sospechoso — redondear a decena más cercana
                # 197006 → 197010? No, redondear a 100 más cercano
                v_redondeado = round(v / 100) * 100
                # Solo aplicar si la diferencia es pequeña (< 20)
                if abs(v - v_redondeado) <= 10:
                    return v_redondeado
        return v
    except ValueError:
        return None


@login_required
def ia_capturar_dummy(request):
    return JsonResponse({
        "success": True,
        "entrada": 150000,
        "salida": 30000,
        "total": 120000,
        "mensaje": "Datos capturados exitosamente (DEMO)",
    })


# ============================================
# TABLAS Y EXCEL
# ============================================

@login_required
def tablas_view(request):
    lecturas = LecturaMaquina.objects.all()

    if request.user.role == "usuario":
        turno_abierto = Turno.objects.filter(usuario=request.user, estado="Abierto").first()
        if turno_abierto:
            lecturas = lecturas.filter(turno=turno_abierto)
        else:
            hoy = timezone.now().date()
            lecturas = lecturas.filter(usuario=request.user, fecha_registro__date=hoy)

    if request.user.role == "admin":
        sucursal_id = request.GET.get("sucursal")
        zona_id = request.GET.get("zona")
        maquina_id = request.GET.get("maquina")
        usuario_id = request.GET.get("usuario")
        fecha_desde = request.GET.get("fecha_desde")
        fecha_hasta = request.GET.get("fecha_hasta")

        hay_filtros = any([sucursal_id, zona_id, maquina_id, usuario_id, fecha_desde, fecha_hasta])

        if sucursal_id:
            lecturas = lecturas.filter(sucursal_id=sucursal_id)
        if zona_id:
            lecturas = lecturas.filter(zona_id=zona_id)
        if maquina_id:
            lecturas = lecturas.filter(maquina_id=maquina_id)
        if usuario_id:
            lecturas = lecturas.filter(usuario_id=usuario_id)
        if fecha_desde:
            lecturas = lecturas.filter(fecha_trabajo__gte=fecha_desde)
        if fecha_hasta:
            lecturas = lecturas.filter(fecha_trabajo__lte=fecha_hasta)

        # Si no hay ningún filtro activo, mostrar solo los registros de HOY por defecto
        if not hay_filtros:
            hoy = timezone.now().date()
            lecturas = lecturas.filter(fecha_trabajo=hoy)

    fecha = request.GET.get("fecha")
    if fecha and request.user.role == "usuario":
        lecturas = lecturas.filter(fecha_trabajo=fecha)

    lecturas = lecturas.select_related("maquina", "zona", "sucursal", "turno", "usuario").order_by("-fecha_trabajo", "-fecha_registro")

    total_entrada = lecturas.aggregate(Sum("entrada"))["entrada__sum"] or 0
    total_salida = lecturas.aggregate(Sum("salida"))["salida__sum"] or 0
    total_total = lecturas.aggregate(Sum("total"))["total__sum"] or 0

    from django.core.paginator import Paginator
    paginator = Paginator(lecturas, 15)
    lecturas_page = paginator.get_page(request.GET.get("page", 1))

    context = {
        "lecturas": lecturas_page,
        "page_obj": lecturas_page,
        "total_entrada": total_entrada,
        "total_salida": total_salida,
        "total_total": total_total,
        "sucursales": Sucursal.objects.filter(is_active=True).order_by("nombre"),
        "zonas": Zona.objects.filter(is_active=True).order_by("sucursal", "orden", "nombre"),
        "maquinas": Maquina.objects.annotate(estado_ord=_estado_ord()).order_by("estado_ord", "numero_maquina"),
        "usuarios": Usuario.objects.filter(role="usuario"),
    }

    return render(request, "tablas.html", context)
class LecturaEditView(UpdateView):
    model = LecturaMaquina
    form_class = LecturaMaquinaForm
    template_name = "control/lectura_edit.html"

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        lectura = self.get_object()
        kwargs["turno"] = lectura.turno
        return kwargs

    def form_valid(self, form):
        response = super().form_valid(form)
        lectura = self.object

        # Sincronizar la línea en ControlLecturas si existe
        control = ControlLecturas.objects.filter(
            sucursal=lectura.sucursal,
            fecha_trabajo=lectura.fecha_trabajo,
        ).first()
        if control:
            ControlLecturasLinea.objects.filter(
                control=control, maquina=lectura.maquina
            ).update(
                entrada_historica=int(lectura.entrada or 0),
                salida_historica=int(lectura.salida or 0),
                entrada_parcial=int(lectura.entrada_dia or 0),
                salida_parcial=int(lectura.salida_dia or 0),
                total=int(lectura.total or 0),
            )
            total_general = control.lineas.aggregate(s=Sum("total"))["s"] or 0
            control.total_general = total_general
            control.save(update_fields=["total_general"])

        return response

    def get_success_url(self):
        lectura = self.object
        control = ControlLecturas.objects.filter(
            sucursal=lectura.sucursal,
            fecha_trabajo=lectura.fecha_trabajo,
        ).first()
        if control:
            return reverse_lazy("control:controles_detail", kwargs={"pk": control.pk})
        return reverse_lazy("control:controles_list")
@login_required
def export_excel(request):
    lecturas = LecturaMaquina.objects.all()

    if request.user.role == "usuario":
        turno_abierto = Turno.objects.filter(usuario=request.user, estado="Abierto").first()
        if turno_abierto:
            lecturas = lecturas.filter(turno=turno_abierto)
        else:
            hoy = timezone.now().date()
            lecturas = lecturas.filter(usuario=request.user, fecha_registro__date=hoy)

    if request.user.role == "admin":
        sucursal_id = request.GET.get("sucursal")
        zona_id = request.GET.get("zona")
        maquina_id = request.GET.get("maquina")
        usuario_id = request.GET.get("usuario")
        fecha_desde = request.GET.get("fecha_desde")
        fecha_hasta = request.GET.get("fecha_hasta")

        if sucursal_id:
            lecturas = lecturas.filter(sucursal_id=sucursal_id)
        if zona_id:
            lecturas = lecturas.filter(zona_id=zona_id)
        if maquina_id:
            lecturas = lecturas.filter(maquina_id=maquina_id)
        if usuario_id:
            lecturas = lecturas.filter(usuario_id=usuario_id)
        if fecha_desde:
            lecturas = lecturas.filter(fecha_registro__date__gte=fecha_desde)
        if fecha_hasta:
            lecturas = lecturas.filter(fecha_registro__date__lte=fecha_hasta)

    fecha = request.GET.get("fecha")
    if fecha and request.user.role == "usuario":
        lecturas = lecturas.filter(fecha_registro__date=fecha)

    lecturas = lecturas.order_by("-fecha_registro")

    wb = Workbook()
    ws = wb.active
    ws.title = "Lecturas de Máquinas"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Fecha", "Sucursal", "Zona", "Nº Máquina", "Juego",
        "Entrada", "Salida", "Total", "Usuario", "Nota"
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_num, lectura in enumerate(lecturas, 2):
        ws.cell(row=row_num, column=1).value = lectura.fecha_registro.strftime("%d/%m/%Y %H:%M")
        ws.cell(row=row_num, column=2).value = lectura.sucursal.nombre if lectura.sucursal else ""
        ws.cell(row=row_num, column=3).value = lectura.zona.nombre if lectura.zona else ""
        ws.cell(row=row_num, column=4).value = lectura.numero_maquina
        ws.cell(row=row_num, column=5).value = lectura.nombre_juego
        ws.cell(row=row_num, column=6).value = lectura.entrada
        ws.cell(row=row_num, column=7).value = lectura.salida
        ws.cell(row=row_num, column=8).value = lectura.total
        ws.cell(row=row_num, column=9).value = getattr(lectura.usuario, "nombre", lectura.usuario.username)
        ws.cell(row=row_num, column=10).value = lectura.nota

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=lecturas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    wb.save(response)
    return response


# ============================================
# CRUD SUCURSALES (ADMIN)
# ============================================

@login_required
@role_required(*ROLES_CONFIG_VER)
def sucursales_list(request):
    query  = request.GET.get("q")
    role   = request.user.role

    sucursales = Sucursal.objects.filter(is_active=True)

    if role in ('supervisor', 'encargado'):
        sucursales = request.user.sucursales.filter(is_active=True)

    if query:
        sucursales = sucursales.filter(nombre__icontains=query)

    from django.core.paginator import Paginator
    paginator = Paginator(sucursales, 15)
    sucursales = paginator.get_page(request.GET.get("page", 1))
    return render(request, "sucursales/list.html", {
        "sucursales":   sucursales,
        "page_obj":     sucursales,
        "query":        query,
        "puede_editar": role in ('admin', 'tecnico'),
    })


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def sucursal_create(request):
    if request.method == "POST":
        form = SucursalForm(request.POST)
        if form.is_valid():
            sucursal = form.save()
            if "guardar_zona" in request.POST:
                messages.success(request, f"Sucursal «{sucursal.nombre}» creada. Ahora agrega sus zonas.")
                return redirect(f"{reverse_lazy('control:zona_create')}?sucursal_id={sucursal.pk}")
            messages.success(request, "Sucursal creada exitosamente.")
            return redirect("control:sucursales_list")
    else:
        form = SucursalForm()

    return render(request, "sucursales/form.html", {"form": form, "title": "Crear Sucursal"})


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def sucursal_edit(request, pk):
    sucursal = get_object_or_404(Sucursal, pk=pk)
    if request.method == "POST":
        form = SucursalForm(request.POST, instance=sucursal)
        if form.is_valid():
            form.save()
            messages.success(request, "Sucursal actualizada exitosamente.")
            return redirect("control:sucursales_list")
    else:
        form = SucursalForm(instance=sucursal)

    return render(request, "sucursales/form.html", {"form": form, "title": "Editar Sucursal"})


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def sucursal_delete(request, pk):
    sucursal = get_object_or_404(Sucursal, pk=pk)

    if request.method == "POST":
        sucursal.is_active = False
        sucursal.save()
        messages.success(request, "Sucursal desactivada correctamente.")
        return redirect("control:sucursales_list")

    return render(request, "sucursales/delete.html", {
        "sucursal": sucursal
    })


# ============================================
# CRUD ZONAS (ADMIN)
# ============================================

@login_required
@role_required(*ROLES_CONFIG_VER)
def zonas_list(request):
    sucursal_id   = request.GET.get("sucursal")
    role          = request.user.role

    zonas         = Zona.objects.filter(is_active=True, sucursal__is_active=True).select_related("sucursal")
    sucursales_qs = Sucursal.objects.filter(is_active=True)

    if role in ('supervisor', 'encargado'):
        mis_suc       = request.user.sucursales.filter(is_active=True)
        zonas         = zonas.filter(sucursal__in=mis_suc)
        sucursales_qs = mis_suc

    if sucursal_id:
        zonas = zonas.filter(sucursal_id=sucursal_id)

    from django.core.paginator import Paginator
    paginator = Paginator(zonas, 15)
    zonas = paginator.get_page(request.GET.get("page", 1))
    return render(request, "zonas/list.html", {
        "zonas":        zonas,
        "page_obj":     zonas,
        "sucursales":   sucursales_qs,
        "sucursal_id":  sucursal_id,
        "puede_editar": role in ('admin', 'tecnico'),
    })


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def zona_create(request):
    sucursal_id = request.GET.get("sucursal_id") or request.POST.get("_sucursal_id")
    if request.method == "POST":
        form = ZonaForm(request.POST)
        if form.is_valid():
            zona = form.save()
            suc_id = zona.sucursal_id
            if "guardar_maquinas" in request.POST:
                # Primera zona de esa sucursal (orden de creación)
                primera_zona = Zona.objects.filter(
                    sucursal_id=suc_id, is_active=True
                ).order_by("pk").first()
                z_id = primera_zona.pk if primera_zona else zona.pk
                messages.success(request, "Zona guardada. Ahora agrega las máquinas.")
                return redirect(
                    f"{reverse_lazy('control:maquina_create')}?sucursal_id={suc_id}&zona_id={z_id}"
                )
            if "guardar_otro" in request.POST:
                messages.success(request, "Zona guardada. Puede agregar otra.")
                return redirect(
                    f"{reverse_lazy('control:zona_create')}?sucursal_id={suc_id}"
                )
            messages.success(request, "Zona creada exitosamente.")
            return redirect("control:zonas_list")
    else:
        initial = {"sucursal": sucursal_id} if sucursal_id else {}
        form = ZonaForm(initial=initial)

    return render(request, "zonas/form.html", {
        "form": form,
        "title": "Crear Zona",
        "sucursal_id": sucursal_id,
    })


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def zona_edit(request, pk):
    zona = get_object_or_404(Zona, pk=pk)
    if request.method == "POST":
        form = ZonaForm(request.POST, instance=zona)
        if form.is_valid():
            form.save()
            messages.success(request, "Zona actualizada exitosamente.")
            return redirect("control:zonas_list")
    else:
        form = ZonaForm(instance=zona)

    return render(request, "zonas/form.html", {"form": form, "title": "Editar Zona"})


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def zona_delete(request, pk):
    zona = get_object_or_404(Zona, pk=pk)

    if request.method == "POST":
        zona.is_active = False
        zona.save()
        Maquina.objects.filter(zona=zona).update(is_active=False)

        messages.success(request, "Zona desactivada correctamente. El histórico fue conservado.")
        return redirect("control:zonas_list")

    return render(request, "zonas/delete.html", {"object": zona})


# ============================================
# CRUD MÁQUINAS (ADMIN)
# ============================================

@login_required
@role_required(*ROLES_CONFIG_VER)
def maquinas_list(request):
    sucursal_id = request.GET.get("sucursal")
    zona_id     = request.GET.get("zona")
    estado      = request.GET.get("estado", "Mantenimiento")
    role        = request.user.role

    maquinas      = Maquina.objects.filter(sucursal__is_active=True, zona__is_active=True).select_related("sucursal", "zona").annotate(estado_ord=_estado_ord()).order_by("estado_ord", "sucursal__nombre", "zona__orden", "numero_maquina")
    sucursales_qs = Sucursal.objects.filter(is_active=True)

    if role in ('supervisor', 'encargado'):
        mis_suc       = request.user.sucursales.filter(is_active=True)
        maquinas      = maquinas.filter(sucursal__in=mis_suc)
        sucursales_qs = mis_suc

    if sucursal_id:
        maquinas = maquinas.filter(sucursal_id=sucursal_id)
    if zona_id:
        maquinas = maquinas.filter(zona_id=zona_id)
    if estado:
        maquinas = maquinas.filter(estado=estado)

    from django.core.paginator import Paginator
    paginator = Paginator(maquinas.order_by("sucursal__nombre", "zona__nombre", "numero_maquina"), 15)
    maquinas = paginator.get_page(request.GET.get("page", 1))
    return render(request, "maquinas/list.html", {
        "maquinas":         maquinas,
        "page_obj":         maquinas,
        "sucursales":       sucursales_qs,
        "zonas":            Zona.objects.filter(is_active=True, sucursal__in=sucursales_qs),
        "sucursal_id":      sucursal_id,
        "zona_id":          zona_id,
        "estado_filtro":    estado,
        "estado_choices":   Maquina.ESTADO_CHOICES,
        "puede_editar":     role in ('admin', 'tecnico'),
    })


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def maquina_create(request):
    sucursal_id = request.GET.get("sucursal_id") or request.POST.get("_sucursal_id")
    zona_id     = request.GET.get("zona_id")     or request.POST.get("_zona_id")
    if request.method == "POST":
        form = MaquinaForm(request.POST)
        if form.is_valid():
            maquina = form.save()
            if "guardar_otro" in request.POST:
                messages.success(request, "Máquina guardada. Puede agregar otra.")
                return redirect(
                    f"{reverse_lazy('control:maquina_create')}?sucursal_id={maquina.sucursal_id}&zona_id={maquina.zona_id}"
                )
            messages.success(request, "Máquinas creadas exitosamente.")
            return redirect("control:maquinas_list")
    else:
        initial = {}
        if sucursal_id:
            initial["sucursal"] = sucursal_id
        if zona_id:
            initial["zona"] = zona_id
        form = MaquinaForm(initial=initial)

    return render(request, "maquinas/form.html", {
        "form": form,
        "title": "Crear Máquina",
        "preselect_sucursal": sucursal_id,
        "preselect_zona": zona_id,
    })


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def maquina_edit(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    if request.method == "POST":
        form = MaquinaForm(request.POST, instance=maquina)
        if form.is_valid():
            form.save()
            messages.success(request, "Máquina actualizada exitosamente.")
            return redirect("control:maquinas_list")
    else:
        form = MaquinaForm(instance=maquina)

    return render(request, "maquinas/form.html", {"form": form, "title": "Editar Máquina"})


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def maquina_delete(request, pk):
    maquina = get_object_or_404(Maquina, pk=pk)
    if request.method == "POST":
        maquina.delete()
        messages.success(request, "Máquina eliminada exitosamente.")
        return redirect("control:maquinas_list")
    return render(request, "maquinas/delete.html", {"object": maquina})


@login_required
@role_required(*ROLES_CONFIG_EDIT)
def maquina_update_estado(request, pk):
    if request.method != "POST":
        return redirect("control:maquinas_list")

    maquina = get_object_or_404(Maquina, pk=pk)

    nuevo_estado = request.POST.get("estado")
    estados_validos = [e[0] for e in Maquina.ESTADO_CHOICES]

    if nuevo_estado not in estados_validos:
        messages.error(request, "Estado inválido.")
        return redirect("control:maquinas_list")

    maquina.estado = nuevo_estado
    maquina.save(update_fields=["estado"])

    messages.success(request, f"Estado actualizado a: {nuevo_estado}")
    return redirect("control:maquinas_list")


# ============================================
# CRUD USUARIOS (ADMIN)
# ============================================

@login_required
@role_required(*ROLES_VER_USUARIOS)
def usuarios_list(request):
    usuarios = Usuario.objects.all().order_by("username").prefetch_related("sucursales")
    from django.core.paginator import Paginator
    paginator = Paginator(usuarios, 15)
    usuarios = paginator.get_page(request.GET.get("page", 1))
    return render(request, "usuarios/list.html", {"usuarios": usuarios, "page_obj": usuarios})


@login_required
@role_required(*ROLES_USUARIOS)
def usuario_create(request):
    if request.method == "POST":
        form = UsuarioForm(request.POST)
        if form.is_valid():
            usuario = form.save(commit=False)
            usuario.set_password(form.cleaned_data["password"])
            usuario.save()
            form.save_m2m()
            messages.success(request, "Usuario creado exitosamente.")
            return redirect("control:usuarios_list")
    else:
        form = UsuarioForm()
    return render(request, "usuarios/form.html", {"form": form, "title": "Crear Usuario"})


@login_required
@role_required(*ROLES_USUARIOS)
def usuario_edit(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == "POST":
        form = UsuarioEditForm(request.POST, instance=usuario)
        if form.is_valid():
            usuario = form.save(commit=False)
            password = form.cleaned_data.get("password")
            if password:
                usuario.set_password(password)
            usuario.save()
            form.save_m2m()
            messages.success(request, "Usuario actualizado exitosamente.")
            return redirect("control:usuarios_list")
    else:
        form = UsuarioEditForm(instance=usuario)
    return render(request, "usuarios/form.html", {"form": form, "title": "Editar Usuario"})


@login_required
@role_required(*ROLES_USUARIOS)
def usuario_delete(request, pk):
    usuario = get_object_or_404(Usuario, pk=pk)
    if request.method == "POST":
        usuario.delete()
        messages.success(request, "Usuario eliminado exitosamente.")
        return redirect("control:usuarios_list")
    return render(request, "usuarios/delete.html", {"object": usuario})

def _seed_cierre_defaults(cierre: CierreTurno):
    # Zonas activas de la sucursal
    zonas = Zona.objects.filter(sucursal=cierre.sucursal, is_active=True).order_by("orden", "nombre")
    for z in zonas:
        CierreTurnoZona.objects.get_or_create(cierre=cierre, zona=z)

    # Movimientos fijos
    for key, _label in CierreTurnoMovimiento.TIPO_CHOICES:
        CierreTurnoMovimiento.objects.get_or_create(cierre=cierre, tipo=key)

    # Pagos fijos
    for key, _label in CierreTurnoPago.TIPO_CHOICES:
        CierreTurnoPago.objects.get_or_create(cierre=cierre, tipo=key)

    # Denominaciones típicas
    billetes = [20000, 10000, 5000, 2000, 1000]
    monedas = [500, 100, 50, 10]
    for d in billetes:
        CierreTurnoDenominacion.objects.get_or_create(cierre=cierre, tipo="BILLETE", denominacion=d)
    for d in monedas:
        CierreTurnoDenominacion.objects.get_or_create(cierre=cierre, tipo="MONEDA", denominacion=d)

from django.db import transaction

# ==========================
# CIERRE TURNO (COLUMNA ROJA)
# ==========================

def _recalcular_totales_cierre(cierre: CierreTurno):
    zonas_qs = cierre.zonas.all()
    movs_qs = cierre.movimientos.all()
    dens_qs = cierre.denominaciones.all()

    total_numeral = sum((x.numeral or 0) for x in zonas_qs)
    total_gastos = sum((m.monto or 0) for m in movs_qs)
    total_efectivo = sum((d.cantidad or 0) for d in dens_qs)
    total_redbank = cierre.pagos.filter(tipo="REDBANK").aggregate(s=Sum("monto"))["s"] or 0

    cierre.total_numeral = int(total_numeral)
    cierre.total_gastos = int(total_gastos)
    cierre.total_efectivo_contado = int(total_efectivo)

    # Si tu modelo tiene campo total_pagos, lo guardas. Si no, lo ignoras.


    cierre.total_esperado = int(
        (cierre.caja_base or 0)
        + total_numeral
        + (cierre.redbank_retiros or 0)      # ✅ se suma
        + (cierre.prestamos_salida or 0)     # ⚠️ por ahora lo sumo; confirma si debe restar
        - (total_gastos + (cierre.retiro_diario or 0))
    )
    cierre.descuadre = int((cierre.total_efectivo_contado or 0) - cierre.total_esperado)

def _autollenar_numerales_por_zona(cierre, turno):
    sums = (
        LecturaMaquina.objects
        .filter(turno=turno)
        .values("zona_id")
        .annotate(total=Sum("total"))
    )

    mapa = {x["zona_id"]: int(x["total"] or 0) for x in sums}

    for cz in cierre.zonas.all():
        nuevo = mapa.get(cz.zona_id, 0)
        if cz.numeral != nuevo:
            cz.numeral = nuevo
            cz.save(update_fields=["numeral"])

def total_entregado_zona(cz):
    return int(
        (cz.billete_20000_monto or 0) +
        (cz.billete_10000_monto or 0) +
        (cz.billete_5000_monto  or 0) +
        (cz.billete_2000_monto  or 0) +
        (cz.billete_1000_monto  or 0) +
        (cz.monedas_monto or 0)
    )

def _recalcular_totales_cierre(cierre: CierreTurno):
    zonas_qs = cierre.zonas.all()
    movs_qs = cierre.movimientos.all()

    total_numeral = sum((z.numeral or 0) for z in zonas_qs)
    total_gastos = sum((m.monto or 0) for m in movs_qs)

    # ✅ AQUÍ estaba el problema: ahora el efectivo viene desde las ZONAS (billetes + monedas)
    total_efectivo = sum(total_entregado_zona(z) for z in zonas_qs)

    cierre.total_numeral = int(total_numeral)
    cierre.total_gastos = int(total_gastos)
    cierre.total_efectivo_contado = int(total_efectivo)

    cierre.total_esperado = int(
        (cierre.caja_base or 0)
        + total_numeral
        + (cierre.redbank_retiros or 0)
        + (cierre.prestamos_salida or 0)
        - (total_gastos + (cierre.retiro_diario or 0))
    )
    cierre.descuadre = int((cierre.total_efectivo_contado or 0) - cierre.total_esperado)


from django.db.models import Sum
from django.utils import timezone

@login_required
def cierre_turno_create_or_edit(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)

    # solo la atendedora del turno o admin
    if request.user.role == "usuario" and turno.usuario_id != request.user.id:
        return HttpResponse("No autorizado", status=403)

    cierre, _created = CierreTurno.objects.get_or_create(
        turno=turno,
        defaults={
            "sucursal": turno.sucursal,
            "usuario": request.user,
            "fecha": getattr(turno, "fecha", None) or timezone.now().date(),
        },
    )

    # --- coherencia cierre vs turno ---
    changed = False
    if cierre.sucursal_id != turno.sucursal_id:
        cierre.sucursal = turno.sucursal
        changed = True

    if cierre.usuario_id != request.user.id:
        cierre.usuario = request.user
        changed = True

    fecha_turno = getattr(turno, "fecha", None) or timezone.now().date()
    if cierre.fecha != fecha_turno:
        cierre.fecha = fecha_turno
        changed = True

    if changed:
        cierre.save()

    # --- sincronizar filas base ---
    _sync_cierre_zonas_activo(cierre)
    _seed_cierre_defaults(cierre)
    _autollenar_numerales_por_zona(cierre, turno)
    _ensure_billetes(cierre)


    for z in cierre.zonas.all():
        total_zona = (
            LecturaMaquina.objects
            .filter(turno=turno, zona=z.zona)
            .aggregate(s=Sum("total"))["s"]
            or 0
        )

        z.numeral = int(total_zona)
        z.save(update_fields=["numeral"])
    # lecturas (para la tabla del template)
    lecturas = (
        LecturaMaquina.objects
        .filter(turno=turno)
        .select_related("zona", "maquina")
        .order_by("zona__orden", "zona__nombre", "numero_maquina", "nombre_juego", "fecha_registro")
    )

    # ✅ SIEMPRE crea form/formsets (GET y POST)
    if request.method == "POST":
        form = CierreTurnoForm(request.POST, instance=cierre)
        zonas_fs = CierreTurnoZonaFormSet(request.POST, instance=cierre, prefix="zonas")
        mov_fs = CierreTurnoMovimientoFormSet(request.POST, instance=cierre, prefix="movimientos")
        pagos_fs = CierreTurnoPagoFormSet(request.POST, instance=cierre, prefix="pagos")
        den_fs = CierreTurnoDenFormSet(request.POST, instance=cierre, prefix="den")

        if form.is_valid() and zonas_fs.is_valid() and mov_fs.is_valid() and pagos_fs.is_valid() and den_fs.is_valid():
            with transaction.atomic():
                form.save()
                zonas_fs.save()
                mov_fs.save()
                pagos_fs.save()
                den_fs.save()
                
                _autollenar_numerales_por_zona(cierre, turno)


                _recalcular_totales_cierre(cierre)
                cierre.save()
                total_lecturas = LecturaMaquina.objects.filter(turno=turno).aggregate(
                    total_sum=Sum("total")
                )["total_sum"] or 0

                cierre.total_lecturas_turno = int(total_lecturas)
                cierre.save(update_fields=["total_lecturas_turno"])

                # cerrar turno
                total_lecturas = LecturaMaquina.objects.filter(turno=turno).aggregate(
                    total_sum=Sum("total")
                )["total_sum"] or 0

                Turno.objects.filter(id=turno.id).update(
                    estado="Cerrado",
                    total_cierre=Decimal(str(total_lecturas)),
                )

            messages.success(request, "Cierre de turno guardado y turno cerrado.")
            return redirect("control:turno")
        else:
            messages.error(request, "Hay errores en el formulario. Revisa los campos.")

    else:
        form = CierreTurnoForm(instance=cierre)
        zonas_fs = CierreTurnoZonaFormSet(instance=cierre, prefix="zonas")
        mov_fs = CierreTurnoMovimientoFormSet(instance=cierre, prefix="movimientos")
        pagos_fs = CierreTurnoPagoFormSet(instance=cierre, prefix="pagos")
        den_fs = CierreTurnoDenFormSet(instance=cierre, prefix="den")


            # --- construir bloques por zona (para template tipo Excel) ---
        # Usamos las zonas del CIERRE (ya sincronizadas) para asegurar que siempre existan
        cierrezonas = (
            cierre.zonas
            .select_related("zona")
            .all()
            .order_by("zona__orden", "zona__nombre")
        )

        zona_blocks = []
        for cz in cierrezonas:
            z = cz.zona

            lects = list(
                LecturaMaquina.objects
                .filter(turno=turno, zona=z)
                .order_by("numero_maquina", "nombre_juego", "fecha_registro")
            )

            total_zona = sum(int(x.total or 0) for x in lects)

            # buscar el formset form correspondiente a esa fila (por PK del inline)
            fz = None
            for _fz in zonas_fs:
                if _fz.instance.pk == cz.pk:
                    fz = _fz
                    break

            zona_blocks.append({
                "zona": z,
                "lecturas": lects,
                "total_zona": total_zona,
                "fz": fz,
            })

    return render(
        request,
        "cuadratura/create_turno.html",
        {
            "turno": turno,
            "cierre": cierre,
            "form": form,
            "zonas_fs": zonas_fs,
            "mov_fs": mov_fs,
            "pagos_fs": pagos_fs,
            "den_fs": den_fs,
            "lecturas": lecturas,
            "zona_blocks": zona_blocks,
        },
   


    )

@login_required
@user_passes_test(is_admin, login_url="/login/")
def cierre_turno_list(request):
    qs = CierreTurno.objects.select_related("turno", "sucursal", "usuario")

    sucursal_id = request.GET.get("sucursal")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    qs = qs.order_by("-fecha", "-id")

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 15)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(
        request,
        "cierre_turno/list.html",
        {
            "cierres": page_obj,
            "page_obj": page_obj,
            "sucursales": Sucursal.objects.filter(is_active=True).order_by("nombre"),
        },
    )


@login_required
@role_required(*ROLES_OPERACIONES)
def cuadratura_zona_list(request):
    user = request.user
    qs = CuadraturaZona.objects.select_related("turno", "turno__sucursal", "turno__usuario", "zona")

    sucursal_id = request.GET.get("sucursal")
    zona_id     = request.GET.get("zona")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    # ── Restricciones por rol ──
    sucursal_activa_id   = request.session.get('sucursal_activa_id')
    sucursal_activa_nombre = ""

    if user.role == 'asistente':
        # Solo su sucursal activa y sus zonas asignadas en el turno del encargado
        if sucursal_activa_id:
            qs = qs.filter(turno__sucursal_id=sucursal_activa_id)
            sucursal_obj = Sucursal.objects.filter(pk=sucursal_activa_id).first()
            sucursal_activa_nombre = sucursal_obj.nombre if sucursal_obj else ""
        turno_enc = Turno.objects.filter(
            sucursal_id=sucursal_activa_id, estado="Abierto", usuario__role="encargado"
        ).first()
        zona_ids = []
        if turno_enc:
            zona_ids = list(AsignacionTurnoZona.objects.filter(
                turno=turno_enc, usuario=user
            ).values_list('zona_id', flat=True))
        if zona_ids:
            qs = qs.filter(zona_id__in=zona_ids)
        zonas = Zona.objects.filter(id__in=zona_ids, is_active=True).order_by("nombre")
        sucursales = []

    elif user.role == 'encargado':
        # Solo sus sucursales
        mis_sucursales = user.sucursales.filter(is_active=True)
        qs = qs.filter(turno__sucursal__in=mis_sucursales)
        sucursales = list(mis_sucursales.order_by("nombre"))
        if sucursal_id:
            qs = qs.filter(turno__sucursal_id=sucursal_id)
        zonas = Zona.objects.filter(
            sucursal__in=mis_sucursales, is_active=True
        ).order_by("nombre")

    else:
        # Admin y otros: ven todo
        sucursales = list(Sucursal.objects.filter(is_active=True).order_by("nombre"))
        zonas = Zona.objects.filter(is_active=True).order_by("nombre")
        if sucursal_id:
            qs = qs.filter(turno__sucursal_id=sucursal_id)

    if zona_id:
        qs = qs.filter(zona_id=zona_id)
    if fecha_desde:
        qs = qs.filter(turno__fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(turno__fecha__lte=fecha_hasta)

    qs = qs.order_by("-turno__fecha", "-id")

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 15)
    page_obj = paginator.get_page(request.GET.get("page", 1))

    return render(
        request,
        "cuadratura_zona/list.html",
        {
            "cuadraturas": page_obj,
            "page_obj": page_obj,
            "sucursales": sucursales,
            "zonas": zonas,
            "sucursal_activa_id": sucursal_activa_id,
            "sucursal_activa_nombre": sucursal_activa_nombre,
        },
    )


@login_required
def cierre_turno_detail(request, pk):
    cierre = get_object_or_404(CierreTurno, pk=pk)

    if request.user.role == "usuario" and cierre.usuario_id != request.user.id:
        return HttpResponse("No autorizado", status=403)

    return render(request, "cierre_turno/detail.html", {"cierre": cierre})



@login_required
def cierre_turno_edit(request, pk):
    cierre = get_object_or_404(CierreTurno, pk=pk)
    turno = cierre.turno  # ✅ importante: ahora existe siempre

    if request.user.role == "usuario" and cierre.usuario_id != request.user.id:
        return HttpResponse("No autorizado", status=403)

    # sincroniza solo zonas activas en ESTE cierre
    _sync_cierre_zonas_activo(cierre)

    # crea filas base de mov/pagos/denominaciones si faltan
    _seed_cierre_defaults(cierre)

    if request.method == "POST":
        form = CierreTurnoForm(request.POST, instance=cierre)
        zonas_fs = CierreTurnoZonaFormSet(request.POST, instance=cierre, prefix="zonas")
        mov_fs = CierreTurnoMovimientoFormSet(request.POST, instance=cierre, prefix="movimientos")
        pagos_fs = CierreTurnoPagoFormSet(request.POST, instance=cierre, prefix="pagos")
        den_fs = CierreTurnoDenFormSet(request.POST, instance=cierre, prefix="den")

        if form.is_valid() and zonas_fs.is_valid() and mov_fs.is_valid() and pagos_fs.is_valid() and den_fs.is_valid():
            with transaction.atomic():
                form.save()
                zonas_fs.save()
                mov_fs.save()
                pagos_fs.save()
                den_fs.save()

                _recalcular_totales_cierre(cierre)
                cierre.save()

                # ✅ cerrar turno si estaba abierto
                if (turno.estado or "").strip().lower() == "abierto":
                    total_lecturas = LecturaMaquina.objects.filter(turno=turno).aggregate(
                        total_sum=Sum("total")
                    )["total_sum"] or 0

                    turno.estado = "Cerrado"
                    turno.total_cierre = Decimal(str(total_lecturas))
                    turno.save(update_fields=["estado", "total_cierre"])

            messages.success(request, "Cierre de turno guardado y turno cerrado.")
            return redirect("control:turno")
        else:
            messages.error(request, "Hay errores en el formulario.")
    else:
        form = CierreTurnoForm(instance=cierre)
        zonas_fs = CierreTurnoZonaFormSet(instance=cierre, prefix="zonas")
        mov_fs = CierreTurnoMovimientoFormSet(instance=cierre, prefix="movimientos")
        pagos_fs = CierreTurnoPagoFormSet(instance=cierre, prefix="pagos")
        den_fs = CierreTurnoDenFormSet(instance=cierre, prefix="den")

    return render(
        request,
        "cuadratura/create_turno.html",
        {
            "turno": turno,
            "cierre": cierre,
            "form": form,
            "zonas_fs": zonas_fs,
            "mov_fs": mov_fs,
            "pagos_fs": pagos_fs,
            "den_fs": den_fs,
        },
    )


@login_required
@user_passes_test(is_admin, login_url="/login/")
def cierre_turno_delete(request, pk):
    cierre = get_object_or_404(CierreTurno, pk=pk)

    if request.method == "POST":
        cierre.delete()
        messages.success(request, "Cierre eliminado.")
        return redirect("control:cierre_turno_list")

    return render(request, "cierre_turno/delete.html", {"cierre": cierre})


@login_required
@user_passes_test(is_admin, login_url="/login/")
def cierre_turno_export_excel(request):
    qs = CierreTurno.objects.select_related("turno", "sucursal", "usuario").order_by("-fecha", "-creado_el")

    sucursal_id = request.GET.get("sucursal")
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")

    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if fecha_desde:
        qs = qs.filter(fecha__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__lte=fecha_hasta)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cierres de Turno"

    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    headers = [
        "Fecha", "Sucursal", "Usuario", "Turno",
        "Caja base", "Retiro", "Total numeral", "Total gastos",
        "Esperado", "Contado", "Descuadre"
    ]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    row = 2
    for x in qs:
        ws.cell(row=row, column=1, value=x.fecha.strftime("%d/%m/%Y") if x.fecha else "")
        ws.cell(row=row, column=2, value=x.sucursal.nombre if x.sucursal else "")
        ws.cell(row=row, column=3, value=getattr(x.usuario, "nombre", x.usuario.username) if x.usuario else "")
        ws.cell(row=row, column=4, value=getattr(x.turno, "tipo_turno", "") if x.turno else "")
        ws.cell(row=row, column=5, value=int(x.caja_base or 0))
        ws.cell(row=row, column=6, value=int(x.retiro_diario or 0))
        ws.cell(row=row, column=7, value=int(x.total_numeral or 0))
        ws.cell(row=row, column=8, value=int(x.total_gastos or 0))
        ws.cell(row=row, column=9, value=int(x.total_esperado or 0))
        ws.cell(row=row, column=10, value=int(x.total_efectivo_contado or 0))
        ws.cell(row=row, column=11, value=int(x.descuadre or 0))
        row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=cierres_turno_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
    wb.save(response)
    return response


from .models import Zona, CierreTurnoZona

def _sync_cierre_zonas_activo(cierre):
    zonas_activas = (
        Zona.objects
        .filter(sucursal=cierre.sucursal, is_active=True)
        .order_by("orden", "nombre")
    )

    zonas_ids = {z.id for z in zonas_activas}

    # borrar del cierre las zonas que ya no están activas
    cierre.zonas.exclude(zona_id__in=zonas_ids).delete()

    # crear filas faltantes para zonas activas
    existentes = set(cierre.zonas.values_list("zona_id", flat=True))
    for z in zonas_activas:
        if z.id not in existentes:
            CierreTurnoZona.objects.create(
                cierre=cierre,
                zona=z,
                numeral=0,
                caja=0,
                prestamos=0,
                redbank=0,
                retiros=0,
                detalle_entregado_total=0,
                descuadre=0,
            )

def _ensure_billetes(cierre):
    billetes = [20000, 10000, 5000, 2000, 1000]
    for d in billetes:
        CierreTurnoDenominacion.objects.get_or_create(
            cierre=cierre,
            tipo="BILLETE",
            denominacion=d,
            defaults={"cantidad": 0}  # OJO: aquí "cantidad" la usas como MONTO ($), no cantidad de billetes
        )


@login_required
@transaction.atomic
def generar_control(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)

    # seguridad: dueño del turno, admin, o asistente asignado a una zona
    es_dueno_gen = (turno.usuario == request.user or request.user.role == "admin")
    es_asignado_gen = AsignacionTurnoZona.objects.filter(turno=turno, usuario=request.user).exists()
    if not es_dueno_gen and not es_asignado_gen:
        return HttpResponse("No autorizado", status=403)

    lecturas = (
        LecturaMaquina.objects
        .filter(turno=turno)
        .select_related("maquina", "maquina__sucursal", "zona")
        .order_by("zona__orden", "zona__nombre", "numero_maquina")
    )

    # Resync: recalcular campos derivados de cualquier lectura que haya sido
    # editada con el modelo viejo (entrada_dia/salida_dia/total desactualizados).
    # Usamos update() directo a la BD para no disparar el save() del modelo
    # (que recalcularía entrada_anterior, lo cual NO queremos cambiar).
    from django.db.models import F
    for lec in lecturas:
        entrada_dia_correcto = int(lec.entrada or 0) - int(lec.entrada_anterior or 0)
        salida_dia_correcto  = int(lec.salida  or 0) - int(lec.salida_anterior  or 0)
        total_correcto       = entrada_dia_correcto - salida_dia_correcto

        # Solo actualizar si hay diferencia (evitar writes innecesarios)
        if (lec.entrada_dia != entrada_dia_correcto or
            lec.salida_dia  != salida_dia_correcto  or
            lec.total       != total_correcto):
            LecturaMaquina.objects.filter(pk=lec.pk).update(
                entrada_dia=entrada_dia_correcto,
                salida_dia=salida_dia_correcto,
                total=total_correcto,
            )

    # Refrescar queryset con valores actualizados
    lecturas = (
        LecturaMaquina.objects
        .filter(turno=turno)
        .select_related("maquina", "zona")
        .order_by("zona__orden", "zona__nombre", "numero_maquina")
    )

    total_entrada_dia = lecturas.aggregate(s=Sum("entrada_dia"))["s"] or 0
    total_salida_dia  = lecturas.aggregate(s=Sum("salida_dia"))["s"] or 0
    total_total       = lecturas.aggregate(s=Sum("total"))["s"] or 0

    # ── Resumen por Servidor ─────────────────────────────────────────────────
    from collections import defaultdict
    serv_map = defaultdict(lambda: {"cant": 0, "entrada": 0, "salida": 0, "total": 0})
    juego_map = defaultdict(lambda: {"cant": 0, "entrada": 0, "salida": 0, "total": 0})
    for l in lecturas:
        srv = l.maquina.servidor or "—"
        jue = l.nombre_juego or "—"
        serv_map[srv]["cant"]    += 1
        serv_map[srv]["entrada"] += l.entrada_dia
        serv_map[srv]["salida"]  += l.salida_dia
        serv_map[srv]["total"]   += l.total
        juego_map[jue]["cant"]    += 1
        juego_map[jue]["entrada"] += l.entrada_dia
        juego_map[jue]["salida"]  += l.salida_dia
        juego_map[jue]["total"]   += l.total

    def add_rtp(d):
        result = []
        for name, v in sorted(d.items()):
            rtp = round(v["salida"] / v["entrada"] * 100, 1) if v["entrada"] > 0 else 0
            result.append({"nombre": name, "rtp": rtp, **v})
        return result

    resumen_servidor = add_rtp(serv_map)
    resumen_juego    = add_rtp(juego_map)

    # ── Resumen Caja ─────────────────────────────────────────────────────────
    asig = AsignacionTurnoZona.objects.filter(turno=turno)
    caja_prestamos = asig.aggregate(s=Sum("prestamo"))["s"] or 0
    caja_retiros   = asig.aggregate(s=Sum("retiros"))["s"] or 0
    caja_redbank   = asig.aggregate(s=Sum("banano"))["s"] or 0

    # Movimientos del cierre (si ya está cerrado)
    movimientos = {}
    try:
        cierre_obj = turno.cierre
        for mov in cierre_obj.movimientos.all():
            movimientos[mov.tipo] = movimientos.get(mov.tipo, 0) + mov.monto
    except Exception:
        cierre_obj = None

    # Billetes malos de cuadratura
    from django.db.models import Q
    billetes_malos = CuadraturaCajaDiaria.objects.filter(turno=turno).aggregate(
        s=Sum("ef_billetes_malos")
    )["s"] or 0

    descuadre_total = CuadraturaCajaDiaria.objects.filter(turno=turno).aggregate(
        s=Sum("descuadre_dia")
    )["s"] or 0

    resumen_caja = {
        "numeral":       total_total,
        "prestamos":     caja_prestamos,
        "retiros":       caja_retiros,
        "redbank":       caja_redbank,
        "transferencias": movimientos.get("TRANSFER", 0),
        "sueldo_extra":  movimientos.get("SUELDO_B", 0),
        "sorteos":       movimientos.get("SORTEOS", 0),
        "gastos":        movimientos.get("GASTOS", 0),
        "regalos":       movimientos.get("REGALOS", 0),
        "jugados":       movimientos.get("JUGADOS", 0),
        "billetes_malos": billetes_malos,
        "descuadre":     descuadre_total,
    }
    resumen_caja["total"] = (
        resumen_caja["numeral"]
        + resumen_caja["prestamos"]
        - resumen_caja["retiros"]
        - resumen_caja["redbank"]
        - resumen_caja["transferencias"]
        - resumen_caja["sueldo_extra"]
        - resumen_caja["sorteos"]
        - resumen_caja["gastos"]
        - resumen_caja["regalos"]
        - resumen_caja["jugados"]
        - resumen_caja["billetes_malos"]
        - resumen_caja["descuadre"]
    )

    volver_a = request.GET.get("from", "list")

    return render(request, "control/control_generado.html", {
        "turno":            turno,
        "lecturas":         lecturas,
        "total_entrada_dia": total_entrada_dia,
        "total_salida_dia":  total_salida_dia,
        "total_total":       total_total,
        "resumen_servidor":  resumen_servidor,
        "resumen_juego":     resumen_juego,
        "resumen_caja":      resumen_caja,
        "volver_a":          volver_a,
    })

@login_required
def cerrar_turno_sin_cuadratura(request, turno_id):
    turno = get_object_or_404(Turno, pk=turno_id)

    # si quieres restringir:
    if request.user.role == "usuario" and turno.usuario_id != request.user.id:
        return HttpResponse("No autorizado", status=403)

    total_lecturas = LecturaMaquina.objects.filter(turno=turno).aggregate(
        s=Sum("total")
    )["s"] or 0

    Turno.objects.filter(id=turno.id).update(
        estado="Cerrado",
        total_cierre=Decimal(str(total_lecturas)),
    )

    messages.success(request, "Turno cerrado.")
    if request.user.role in ('encargado', 'asistente'):
        logout(request)
        return redirect("control:login")

    return redirect("control:turno")


from .models import (
    Turno, LecturaMaquina, ControlLecturas, ControlLecturasLinea
)

@login_required
@transaction.atomic
def guardar_control(request, turno_id):
    if request.method != "POST":
        return redirect("control:generar_control", turno_id=turno_id)

    turno = get_object_or_404(Turno, id=turno_id)

    # Permitir: dueño del turno (admin/encargado) O asistente asignado a una zona del turno
    es_dueno = (turno.usuario == request.user or request.user.role == "admin")
    es_asignado = AsignacionTurnoZona.objects.filter(turno=turno, usuario=request.user).exists()
    if not es_dueno and not es_asignado:
        messages.error(request, "No autorizado.")
        return redirect("control:dashboard")

    fecha    = turno.fecha
    sucursal = turno.sucursal

    # Determinar qué zonas le corresponden al usuario actual
    # - Admin/dueño del turno: todas las zonas con lecturas
    # - Asistente: solo las zonas que tiene asignadas en este turno
    if es_dueno:
        lecturas = (
            LecturaMaquina.objects
            .filter(turno=turno)
            .select_related("maquina", "zona")
            .order_by("zona__nombre", "numero_maquina", "id")
        )
        zonas_del_usuario = None  # todas
    else:
        zonas_asignadas = AsignacionTurnoZona.objects.filter(
            turno=turno, usuario=request.user
        ).values_list("zona_id", flat=True)
        lecturas = (
            LecturaMaquina.objects
            .filter(turno=turno, zona_id__in=zonas_asignadas)
            .select_related("maquina", "zona")
            .order_by("zona__nombre", "numero_maquina", "id")
        )
        zonas_del_usuario = list(zonas_asignadas)

    if not lecturas.exists():
        messages.error(request, "No hay lecturas para guardar control.")
        return redirect("control:generar_control", turno_id=turno_id)

    # Crear el control si no existe (el primero que guarda lo crea)
    control, created = ControlLecturas.objects.update_or_create(
        sucursal=sucursal,
        fecha_trabajo=fecha,
        defaults={
            "turno": turno,
            "creado_por": request.user,
        }
    )

    # Borrar SOLO las líneas de las zonas del usuario actual, no tocar las demás
    if zonas_del_usuario is not None:
        control.lineas.filter(zona_id__in=zonas_del_usuario).delete()
    else:
        control.lineas.all().delete()

    bulk = []
    for lec in lecturas:
        servidor = getattr(lec.maquina, "servidor", "") or "" if lec.maquina_id else ""
        bulk.append(ControlLecturasLinea(
            control=control,
            zona=lec.zona,
            maquina=lec.maquina,
            numero_maquina=lec.numero_maquina,
            servidor=servidor,
            juego=lec.nombre_juego,
            entrada_historica=int(lec.entrada or 0),
            salida_historica=int(lec.salida or 0),
            entrada_parcial=int(lec.entrada_dia or 0),
            salida_parcial=int(lec.salida_dia or 0),
            total=int(lec.total or 0),
        ))

    ControlLecturasLinea.objects.bulk_create(bulk)

    # Recalcular total_general sumando TODAS las líneas (de todas las zonas)
    total_general = control.lineas.aggregate(s=Sum("total"))["s"] or 0
    control.total_general = total_general
    control.save(update_fields=["total_general"])

    messages.success(request, "Control guardado correctamente.")
    volver_a = request.POST.get("from", "list")
    return redirect(f"{reverse('control:controles_detail', args=[control.pk])}?from={volver_a}")


@login_required
@login_required
def lectura_edit_ajax(request, pk):
    """Editar entrada/salida de una lectura via AJAX desde el control generado."""
    import json
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método no permitido"}, status=405)

    lectura = get_object_or_404(LecturaMaquina, pk=pk)

    es_dueno = (lectura.turno.usuario == request.user or request.user.role == "admin")
    es_asignado = AsignacionTurnoZona.objects.filter(turno=lectura.turno, usuario=request.user).exists()
    if not es_dueno and not es_asignado:
        return JsonResponse({"ok": False, "error": "No autorizado"}, status=403)

    try:
        data    = json.loads(request.body)
        entrada = int(data.get("entrada", 0))
        salida  = int(data.get("salida",  0))
        nota    = str(data.get("nota",    ""))[:255]
    except (ValueError, KeyError):
        return JsonResponse({"ok": False, "error": "Datos inválidos"}, status=400)

    entrada_dia = entrada - int(lectura.entrada_anterior or 0)
    salida_dia  = salida  - int(lectura.salida_anterior  or 0)
    total       = entrada_dia - salida_dia

    try:
        LecturaMaquina.objects.filter(pk=pk).update(
            entrada=entrada,
            salida=salida,
            entrada_dia=entrada_dia,
            salida_dia=salida_dia,
            total=total,
            nota=nota,
        )
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"Error al guardar lectura: {e}"}, status=500)

    control = ControlLecturas.objects.filter(
        sucursal=lectura.sucursal,
        fecha_trabajo=lectura.fecha_trabajo,
    ).first()
    if control:
        ControlLecturasLinea.objects.filter(
            control=control, maquina=lectura.maquina
        ).update(
            entrada_historica=entrada,
            salida_historica=salida,
            entrada_parcial=entrada_dia,
            salida_parcial=salida_dia,
            total=total,
        )
        total_general = control.lineas.aggregate(s=Sum("total"))["s"] or 0
        control.total_general = total_general
        control.save(update_fields=["total_general"])

    return JsonResponse({
        "ok": True,
        "entrada": entrada,
        "salida": salida,
        "entrada_dia": entrada_dia,
        "salida_dia": salida_dia,
        "total": total,
    })


def lectura_edit_from_control(request, linea_pk, control_pk):
    """Editar una LecturaMaquina accediendo desde el detalle del control."""
    from django.shortcuts import get_object_or_404
    linea   = get_object_or_404(ControlLecturasLinea, pk=linea_pk)
    control = get_object_or_404(ControlLecturas, pk=control_pk)

    # Buscar la LecturaMaquina correspondiente
    lectura = LecturaMaquina.objects.filter(
        maquina=linea.maquina,
        fecha_trabajo=control.fecha_trabajo,
        sucursal=control.sucursal,
    ).first()

    if not lectura:
        messages.error(request, "No se encontró la lectura original para editar.")
        return redirect("control:controles_detail", pk=control_pk)

    from .forms import LecturaMaquinaForm as _Form
    if request.method == "POST":
        form = _Form(request.POST, instance=lectura, turno=lectura.turno)
        if form.is_valid():
            form.save()
            lectura.refresh_from_db()  # obtener campos recalculados (entrada_dia, salida_dia, total)
            # Sync control line
            linea.entrada_historica = int(lectura.entrada or 0)
            linea.salida_historica  = int(lectura.salida or 0)
            linea.entrada_parcial   = int(lectura.entrada_dia or 0)
            linea.salida_parcial    = int(lectura.salida_dia or 0)
            linea.total             = int(lectura.total or 0)
            linea.save()
            total_general = control.lineas.aggregate(s=Sum("total"))["s"] or 0
            control.total_general = total_general
            control.save(update_fields=["total_general"])
            messages.success(request, f"Lectura de máquina {lectura.numero_maquina} actualizada.")
            return redirect("control:controles_detail", pk=control_pk)
    else:
        form = _Form(instance=lectura, turno=lectura.turno)

    return render(request, "control/lectura_edit.html", {
        "form": form,
        "control": control,
        "desde_control": True,
    })

@login_required
@role_required(*ROLES_CONTROLES)
def controles_list(request):
    qs = (
        ControlLecturas.objects
        .select_related("sucursal", "turno", "creado_por")
        .order_by("-fecha_trabajo", "-id")
    )

    sucursal_id  = request.GET.get("sucursal")
    fecha_desde  = request.GET.get("fecha_desde")
    fecha_hasta  = request.GET.get("fecha_hasta")

    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if fecha_desde:
        qs = qs.filter(fecha_trabajo__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_trabajo__lte=fecha_hasta)

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 15)
    controles_page = paginator.get_page(request.GET.get("page", 1))
    return render(request, "controles/list.html", {
        "controles": controles_page,
        "page_obj":  controles_page,
        "sucursales": Sucursal.objects.filter(is_active=True).order_by("nombre"),
    })


@login_required
def controles_detail(request, pk):
    from collections import defaultdict
    control = (
        ControlLecturas.objects
        .select_related("sucursal", "turno", "creado_por")
        .prefetch_related("lineas")
        .get(pk=pk)
    )
    lineas = list(control.lineas.select_related("zona", "maquina").order_by("zona__orden", "zona__nombre", "numero_maquina"))

    total_entrada_parcial = sum(l.entrada_parcial or 0 for l in lineas)
    total_salida_parcial  = sum(l.salida_parcial  or 0 for l in lineas)

    # Resumen por servidor y juego
    serv_map  = defaultdict(lambda: {"cant": 0, "entrada": 0, "salida": 0, "total": 0})
    juego_map = defaultdict(lambda: {"cant": 0, "entrada": 0, "salida": 0, "total": 0})
    for l in lineas:
        srv = l.servidor or "—"
        jue = l.juego    or "—"
        for m, key in ((serv_map, srv), (juego_map, jue)):
            m[key]["cant"]    += 1
            m[key]["entrada"] += l.entrada_parcial or 0
            m[key]["salida"]  += l.salida_parcial  or 0
            m[key]["total"]   += l.total           or 0

    def add_rtp(d):
        result = []
        for name, v in sorted(d.items()):
            rtp = round(v["salida"] / v["entrada"] * 100, 1) if v["entrada"] > 0 else 0
            result.append({"nombre": name, "rtp": rtp, **v})
        return result

    resumen_servidor = add_rtp(serv_map)
    resumen_juego    = add_rtp(juego_map)

    # Resumen caja desde el turno asociado
    movimientos = {}
    caja_prestamos = caja_retiros = caja_redbank = 0
    billetes_malos = descuadre_total = 0
    if control.turno_id:
        try:
            for mov in control.turno.cierre.movimientos.all():
                movimientos[mov.tipo] = movimientos.get(mov.tipo, 0) + mov.monto
        except Exception:
            pass
        asig = AsignacionTurnoZona.objects.filter(turno_id=control.turno_id)
        caja_prestamos = asig.aggregate(s=Sum("prestamo"))["s"] or 0
        caja_retiros   = asig.aggregate(s=Sum("retiros"))["s"]  or 0
        caja_redbank   = asig.aggregate(s=Sum("banano"))["s"]   or 0
        billetes_malos = CuadraturaCajaDiaria.objects.filter(turno_id=control.turno_id).aggregate(s=Sum("ef_billetes_malos"))["s"] or 0
        descuadre_total = CuadraturaCajaDiaria.objects.filter(turno_id=control.turno_id).aggregate(s=Sum("descuadre_dia"))["s"] or 0

    resumen_caja = {
        "numeral":        control.total_general,
        "prestamos":      caja_prestamos,
        "retiros":        caja_retiros,
        "redbank":        caja_redbank,
        "transferencias": movimientos.get("TRANSFER", 0),
        "sueldo_extra":   movimientos.get("SUELDO_B", 0),
        "sorteos":        movimientos.get("SORTEOS", 0),
        "gastos":         movimientos.get("GASTOS", 0),
        "regalos":        movimientos.get("REGALOS", 0),
        "jugados":        movimientos.get("JUGADOS", 0),
        "billetes_malos": billetes_malos,
        "descuadre":      descuadre_total,
    }
    resumen_caja["total"] = (
        resumen_caja["numeral"]
        + resumen_caja["prestamos"]
        - resumen_caja["retiros"]
        - resumen_caja["redbank"]
        - resumen_caja["transferencias"]
        - resumen_caja["sueldo_extra"]
        - resumen_caja["sorteos"]
        - resumen_caja["gastos"]
        - resumen_caja["regalos"]
        - resumen_caja["jugados"]
        - resumen_caja["billetes_malos"]
        - resumen_caja["descuadre"]
    )

    # Navegación anterior / siguiente dentro del mismo local
    _base = ControlLecturas.objects.filter(sucursal=control.sucursal).exclude(pk=control.pk)
    anterior = (
        _base
        .filter(
            Q(fecha_trabajo__lt=control.fecha_trabajo) |
            Q(fecha_trabajo=control.fecha_trabajo, id__lt=control.pk)
        )
        .order_by("-fecha_trabajo", "-id")
        .only("id", "fecha_trabajo")
        .first()
    )
    siguiente = (
        _base
        .filter(
            Q(fecha_trabajo__gt=control.fecha_trabajo) |
            Q(fecha_trabajo=control.fecha_trabajo, id__gt=control.pk)
        )
        .order_by("fecha_trabajo", "id")
        .only("id", "fecha_trabajo")
        .first()
    )

    volver_a = request.GET.get("from", "list")  # "list" | "dashboard"

    return render(request, "controles/detail.html", {
        "control":               control,
        "lineas":                lineas,
        "total_entrada_parcial": total_entrada_parcial,
        "total_salida_parcial":  total_salida_parcial,
        "resumen_servidor":      resumen_servidor,
        "resumen_juego":         resumen_juego,
        "resumen_caja":          resumen_caja,
        "anterior":              anterior,
        "siguiente":             siguiente,
        "volver_a":              volver_a,
    })

@login_required
def turno_asistente_redirect(request):
    turno = Turno.objects.filter(usuario=request.user, estado="Abierto").first()
    if turno:
        return redirect("control:cuadratura_zona", turno_id=turno.id)
    return redirect("control:turno")

# ============================================
# REGISTRO DE ACTIVIDAD — Audit log
# ============================================

@login_required
@role_required('admin', 'gerente')
def movimientos_list(request):
    from django.core.paginator import Paginator
    from .models import RegistroActividad

    qs = RegistroActividad.objects.select_related('usuario', 'sucursal').order_by('-fecha_hora')

    # ── Filtros ────────────────────────────────────────────────
    tipo_fil        = request.GET.get('tipo', '')
    usuario_id      = request.GET.get('usuario', '')
    sucursal_id     = request.GET.get('sucursal', '')
    fecha_desde     = request.GET.get('fecha_desde', '')
    fecha_hasta     = request.GET.get('fecha_hasta', '')
    modulo_fil      = request.GET.get('modulo', '')
    buscar          = request.GET.get('buscar', '')

    if tipo_fil:
        qs = qs.filter(tipo=tipo_fil)
    if usuario_id:
        qs = qs.filter(usuario_id=usuario_id)
    if sucursal_id:
        qs = qs.filter(sucursal_id=sucursal_id)
    if fecha_desde:
        qs = qs.filter(fecha_hora__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_hora__date__lte=fecha_hasta)
    if modulo_fil:
        qs = qs.filter(modulo=modulo_fil)
    if buscar:
        qs = qs.filter(
            Q(nombre_usuario__icontains=buscar) |
            Q(objeto_str__icontains=buscar)     |
            Q(descripcion__icontains=buscar)
        )

    paginator = Paginator(qs, 15)
    page_obj  = paginator.get_page(request.GET.get('page', 1))

    modulos_disponibles = (
        RegistroActividad.objects
        .values_list('modulo', flat=True)
        .distinct()
        .order_by('modulo')
    )

    return render(request, 'movimientos/list.html', {
        'page_obj':   page_obj,
        'tipos':      RegistroActividad.TIPO_CHOICES,
        'usuarios':   Usuario.objects.filter(is_active=True).order_by('nombre'),
        'sucursales': Sucursal.objects.filter(is_active=True).order_by('nombre'),
        'modulos':    modulos_disponibles,
        'filtros': {
            'tipo':        tipo_fil,
            'usuario':     usuario_id,
            'sucursal':    sucursal_id,
            'fecha_desde': fecha_desde,
            'fecha_hasta': fecha_hasta,
            'modulo':      modulo_fil,
            'buscar':      buscar,
        },
    })


# ============================================
# SESIONES ADMIN
# ============================================

@login_required
def sesiones_admin(request):
    """Vista exclusiva del admin para ver todos los registros de inicio y cierre de sesión."""
    if request.user.role != "admin":
        return HttpResponse("No autorizado", status=403)

    filtro_usuario = request.GET.get("usuario", "")
    filtro_rol     = request.GET.get("rol", "")
    filtro_fecha   = request.GET.get("fecha", "")
    filtro_estado  = request.GET.get("estado", "")

    qs = RegistroSesion.objects.select_related("usuario", "sucursal", "turno").order_by("-hora_inicio")

    if filtro_usuario:
        qs = qs.filter(usuario__username__icontains=filtro_usuario)
    if filtro_rol:
        qs = qs.filter(tipo_usuario=filtro_rol)
    if filtro_fecha:
        qs = qs.filter(fecha=filtro_fecha)
    if filtro_estado == "activa":
        qs = qs.filter(hora_cierre__isnull=True)
    elif filtro_estado == "cerrada":
        qs = qs.filter(hora_cierre__isnull=False)

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 15)
    page = request.GET.get("page", 1)
    sesiones = paginator.get_page(page)

    hoy = timezone.localdate()
    activas_hoy  = RegistroSesion.objects.filter(fecha=hoy, hora_cierre__isnull=True).count()
    cerradas_hoy = RegistroSesion.objects.filter(fecha=hoy, hora_cierre__isnull=False).count()
    total_hoy    = RegistroSesion.objects.filter(fecha=hoy).count()

    return render(request, "control/sesiones_admin.html", {
        "sesiones":          sesiones,
        "page_obj":          sesiones,
        "filtro_usuario":    filtro_usuario,
        "filtro_rol":        filtro_rol,
        "filtro_fecha":      filtro_fecha,
        "filtro_estado":     filtro_estado,
        "roles_disponibles": RegistroSesion.TIPO_CHOICES,
        "activas_hoy":       activas_hoy,
        "cerradas_hoy":      cerradas_hoy,
        "total_hoy":         total_hoy,
    })